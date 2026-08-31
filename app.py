from __future__ import annotations

import json
import re
import shutil
import traceback
import unicodedata
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import column_index_from_string, get_column_letter

# ==========================================================
# PATHS
# ==========================================================
BASE_DIR = Path(__file__).resolve().parent
ASSETS_DIR = BASE_DIR / "Assets"
LOGO_PATH = ASSETS_DIR / "Bata_logo.png"

NORMALIZATION_DIR = BASE_DIR / "Normalizacion"
ORIGINALS_DIR = NORMALIZATION_DIR / "Category Originales"
LOCAL_DIR = NORMALIZATION_DIR / "Category ML"
USD_DIR = NORMALIZATION_DIR / "Category USD"

GLOBAL_XLSX = NORMALIZATION_DIR / "Category Global.xlsx"
GLOBAL_DATA_CSV = NORMALIZATION_DIR / "Category Global_data.csv"
GLOBAL_LOCAL_DATA_CSV = NORMALIZATION_DIR / "Category Global_local_data.csv"
UPLOAD_META_PATH = NORMALIZATION_DIR / "upload_metadata.json"

for path in (NORMALIZATION_DIR, ORIGINALS_DIR, LOCAL_DIR, USD_DIR):
    path.mkdir(parents=True, exist_ok=True)

# ==========================================================
# CONSTANTS - SOURCE STRUCTURE
# ==========================================================
SOURCE_SHEET_NAME = "Cat.Plan Template H2'27"
GLOBAL_SHEET_NAME = "Global Cat.Plan H2'27"
METADATA_SHEET_NAME = "Metadata"

SOURCE_START_ROW = 8
SOURCE_MAPPING_ROW = 2
SOURCE_HEADER_ROW = 3
SOURCE_DATA_START_ROW = 4

COL_BU = 2
COL_GENDER = 3
COL_OCCASION = 4
COL_CATEGORY = 5
COL_COUNTRY = column_index_from_string("CP")

# ==========================================================
# HEADERS - EXACT ORDER
# ==========================================================
HEADERS = [
    "KEY", "BU", "GENDER", "OCCASION", "CATEGORY",
    "TURNOVER VALUE 2025", "TURNOVER VALUE 2026", "TURNOVER VALUE 2027",
    "TO NEW COLLECTION", "SHARE %",
    "TO TALY% VS 2025", "TO TALY% VS 2026",
    "TO CONTRIBUTION % 2025", "TO CONTRIBUTION % 2026", "TO CONTRIBUTION % 2027",
    "VACIA",
    "INTAKE% 2025", "INTAKE% 2026", "INTAKE% 2027",
    "INTAKE TALY% VS 2025", "INTAKE TALY% VS 2026",
    "VACIA2",
    "MKD% 2025", "MKD% 2026", "MKD% 2027",
    "VACIA3",
    "MRG% 2025", "MRG% 2026", "MRG% 2027",
    "VACIA4",
    "RRP 2025", "RRP 2026", "RRP 2027",
    "RRP TALY% VS 2025", "RRP TALY% VS 2026",
    "VACIA5",
    "RRP VAT 2025", "RRP VAT 2026", "RRP VAT 2027",
    "RRP VAT TALY% VS 2025", "RRP VAT TALY% VS 2026",
    "VACIA6",
    "ASP 2025", "ASP 2026", "ASP 2027",
    "ASP TALY% VS 2025", "ASP TALY% VS 2026",
    "VACIA7",
    "MKD VALUE 2025", "MKD VALUE 2026", "MKD VALUE 2027",
    "VACIA8",
    "MRG VALUE 2025", "MRG VALUE 2026", "MRG VALUE 2027",
    "VACIA9",
    "COGS VALUE 2025", "COGS VALUE 2026", "COGS VALUE 2027",
    "VACIA10",
    "GROSS VALUE 2025", "GROSS VALUE 2026", "GROSS VALUE 2027",
    "VACIA11",
    "PAIRS VALUE 2025", "PAIRS VALUE 2026", "PAIRS VALUE 2027",
    "PAIRS TALY% VS 2025", "PAIRS TALY% VS 2026",
    "PAIRS CONTRIBUTION % 2025", "PAIRS CONTRIBUTION % 2026", "PAIRS CONTRIBUTION % 2027",
    "VACIA12",
    "INITIAL STOCK QTY VALUE 2025 H1", "FINAL STOCK QTY VALUE 2025 H1", "% FINAL STOCK VS INITIAL 25H1",
    "INITIAL STOCK QTY VALUE 2026 H1", "FINAL STOCK QTY VALUE 2026 H1", "% FINAL STOCK VS INITIAL 26H1",
    "INITIAL STOCK QTY VALUE 2027 H1", "FINAL STOCK QTY VALUE 2027 H1", "% FINAL STOCK VS INITIAL 27H1",
    "VACIA13",
    "OTB QTY 2026 H1", "OTB COST 2026 H1", "SOR FINAL 2026 H2",
    "OTB QTY 2027 H1", "OTB COST 2027 H1", "SOR TARGET 2027H1",
    "VACIA14",
    "STOCK TURN 2025", "STOCK TURN 2026", "STOCK TURN 2027",
    "COUNTRY",
]
HEADER_TO_COL = {header: idx + 1 for idx, header in enumerate(HEADERS)}

# ==========================================================
# ROW 2 - TYPE METADATA
# ==========================================================
ROW2_TYPES = [
    "TEXTO", "TEXTO", "TEXTO", "TEXTO", "TEXTO",
    "VALUE", "VALUE", "VALUE",
    "FORMULA", "FORMULA", "FORMULA", "FORMULA", "FORMULA", "FORMULA", "FORMULA",
    "VACIA",
    "FORMULA", "FORMULA", "FORMULA", "FORMULA", "FORMULA",
    "VACIA2",
    "FORMULA", "FORMULA", "FORMULA",
    "VACIA3",
    "FORMULA", "FORMULA", "FORMULA",
    "VACIA4",
    "FORMULA", "FORMULA", "FORMULA", "FORMULA", "FORMULA",
    "VACIA5",
    "FORMULA", "FORMULA", "FORMULA", "FORMULA", "FORMULA",
    "VACIA6",
    "FORMULA", "FORMULA", "FORMULA", "FORMULA", "FORMULA",
    "VACIA7",
    "VALUE", "VALUE", "VALUE",
    "VACIA8",
    "VALUE", "VALUE", "VALUE",
    "VACIA9",
    "VALUE", "VALUE", "VALUE",
    "VACIA10",
    "VALUE", "VALUE", "VALUE",
    "VACIA11",
    "VALUE PAIRS", "VALUE PAIRS", "VALUE PAIRS",
    "FORMULA", "FORMULA", "FORMULA", "FORMULA", "FORMULA",
    "VACIA12",
    "VALUE PAIRS", "VALUE PAIRS", "FORMULA",
    "VALUE PAIRS", "VALUE PAIRS", "FORMULA",
    "VALUE PAIRS", "VALUE PAIRS", "FORMULA",
    "VACIA13",
    "VALUE PAIRS", "VALUE", "FORMULA",
    "VALUE PAIRS", "VALUE", "FORMULA",
    "VACIA14",
    "FORMULA", "FORMULA", "FORMULA",
    "COUNTRY",
]

# ==========================================================
# EXCLUDED COLUMNS - PRESENT IN STRUCTURE BUT OMITTED FROM UI
# ==========================================================
OMIT_COLUMNS = {"TO NEW COLLECTION", "SHARE %"}

# ==========================================================
# MONEY AND PAIRS COLUMNS
# ==========================================================
MONEY_COLUMNS = [
    "TURNOVER VALUE 2025", "TURNOVER VALUE 2026", "TURNOVER VALUE 2027",
    "MKD VALUE 2025", "MKD VALUE 2026", "MKD VALUE 2027",
    "MRG VALUE 2025", "MRG VALUE 2026", "MRG VALUE 2027",
    "COGS VALUE 2025", "COGS VALUE 2026", "COGS VALUE 2027",
    "GROSS VALUE 2025", "GROSS VALUE 2026", "GROSS VALUE 2027",
    "OTB COST 2026 H1", "OTB COST 2027 H1",
]
PAIRS_SCALABLE_COLUMNS = [
    "PAIRS VALUE 2025", "PAIRS VALUE 2026", "PAIRS VALUE 2027",
    "INITIAL STOCK QTY VALUE 2025 H1", "FINAL STOCK QTY VALUE 2025 H1",
    "INITIAL STOCK QTY VALUE 2026 H1", "FINAL STOCK QTY VALUE 2026 H1",
    "INITIAL STOCK QTY VALUE 2027 H1", "FINAL STOCK QTY VALUE 2027 H1",
    "OTB QTY 2026 H1",
    "OTB QTY 2027 H1",
]
BASE_ADDITIVE_COLUMNS = [
    "TURNOVER VALUE 2025", "TURNOVER VALUE 2026", "TURNOVER VALUE 2027",
    "MKD VALUE 2025", "MKD VALUE 2026", "MKD VALUE 2027",
    "MRG VALUE 2025", "MRG VALUE 2026", "MRG VALUE 2027",
    "COGS VALUE 2025", "COGS VALUE 2026", "COGS VALUE 2027",
    "GROSS VALUE 2025", "GROSS VALUE 2026", "GROSS VALUE 2027",
    "PAIRS VALUE 2025", "PAIRS VALUE 2026", "PAIRS VALUE 2027",
    "INITIAL STOCK QTY VALUE 2025 H1", "FINAL STOCK QTY VALUE 2025 H1",
    "INITIAL STOCK QTY VALUE 2026 H1", "FINAL STOCK QTY VALUE 2026 H1",
    "INITIAL STOCK QTY VALUE 2027 H1", "FINAL STOCK QTY VALUE 2027 H1",
    "OTB QTY 2026 H1", "OTB COST 2026 H1",
    "OTB QTY 2027 H1", "OTB COST 2027 H1",
]
PRESERVED_NON_ADDITIVE = ["TO NEW COLLECTION", "SHARE %"]
TEXT_COLUMNS = ["KEY", "BU", "GENDER", "OCCASION", "CATEGORY", "COUNTRY"]

SOURCE_BASE_MAP = {
    "F": "TURNOVER VALUE 2025",
    "G": "TURNOVER VALUE 2026",
    "H": "TURNOVER VALUE 2027",
    "I": "TO NEW COLLECTION",
    "J": "SHARE %",
    "AW": "MKD VALUE 2025",
    "AX": "MKD VALUE 2026",
    "AY": "MKD VALUE 2027",
    "BA": "MRG VALUE 2025",
    "BB": "MRG VALUE 2026",
    "BC": "MRG VALUE 2027",
    "BE": "COGS VALUE 2025",
    "BF": "COGS VALUE 2026",
    "BG": "COGS VALUE 2027",
    "BI": "GROSS VALUE 2025",
    "BJ": "GROSS VALUE 2026",
    "BK": "GROSS VALUE 2027",
    "BM": "PAIRS VALUE 2025",
    "BN": "PAIRS VALUE 2026",
    "BO": "PAIRS VALUE 2027",
    "BV": "INITIAL STOCK QTY VALUE 2025 H1",
    "BW": "FINAL STOCK QTY VALUE 2025 H1",
    "BY": "INITIAL STOCK QTY VALUE 2026 H1",
    "BZ": "FINAL STOCK QTY VALUE 2026 H1",
    "CB": "INITIAL STOCK QTY VALUE 2027 H1",
    "CC": "FINAL STOCK QTY VALUE 2027 H1",
    "CF": "OTB QTY 2026 H1",
    "CG": "OTB COST 2026 H1",
    "CI": "OTB QTY 2027 H1",
    "CJ": "OTB COST 2027 H1",
}

# ==========================================================
# TABLE DISPLAY ORDERS
# ==========================================================
SEP = None
DEFAULT_TABLE_ORDER = [
    "KEY", "BU", "GENDER", "OCCASION", "CATEGORY",
    "TURNOVER VALUE 2026", "TURNOVER VALUE 2027", "TO TALY% VS 2026",
    "TO CONTRIBUTION % 2026", "TO CONTRIBUTION % 2027",
    SEP,
    "INTAKE% 2026", "INTAKE% 2027", "INTAKE TALY% VS 2026",
    SEP,
    "MKD% 2026", "MKD% 2027",
    SEP,
    "MRG% 2026", "MRG% 2027",
    SEP,
    "RRP VAT 2026", "RRP VAT 2027", "RRP VAT TALY% VS 2026",
    SEP,
    "ASP 2026", "ASP 2027", "ASP TALY% VS 2026",
    "PAIRS VALUE 2026", "PAIRS VALUE 2027", "PAIRS TALY% VS 2026",
    "PAIRS CONTRIBUTION % 2026", "PAIRS CONTRIBUTION % 2027",
]
FULL_TABLE_ORDER = [
    "KEY", "BU", "GENDER", "OCCASION", "CATEGORY",
    "TURNOVER VALUE 2025", "TURNOVER VALUE 2026", "TURNOVER VALUE 2027",
    "TO TALY% VS 2025", "TO TALY% VS 2026",
    "TO CONTRIBUTION % 2025", "TO CONTRIBUTION % 2026", "TO CONTRIBUTION % 2027",
    SEP,
    "INTAKE% 2025", "INTAKE% 2026", "INTAKE% 2027",
    "INTAKE TALY% VS 2025", "INTAKE TALY% VS 2026",
    SEP,
    "MKD% 2025", "MKD% 2026", "MKD% 2027",
    SEP,
    "MRG% 2025", "MRG% 2026", "MRG% 2027",
    SEP,
    "RRP 2025", "RRP 2026", "RRP 2027",
    "RRP TALY% VS 2025", "RRP TALY% VS 2026",
    SEP,
    "RRP VAT 2025", "RRP VAT 2026", "RRP VAT 2027",
    "RRP VAT TALY% VS 2025", "RRP VAT TALY% VS 2026",
    SEP,
    "ASP 2025", "ASP 2026", "ASP 2027",
    "ASP TALY% VS 2025", "ASP TALY% VS 2026",
    SEP,
    "MKD VALUE 2025", "MKD VALUE 2026", "MKD VALUE 2027",
    SEP,
    "MRG VALUE 2025", "MRG VALUE 2026", "MRG VALUE 2027",
    SEP,
    "COGS VALUE 2025", "COGS VALUE 2026", "COGS VALUE 2027",
    SEP,
    "GROSS VALUE 2025", "GROSS VALUE 2026", "GROSS VALUE 2027",
    SEP,
    "PAIRS VALUE 2025", "PAIRS VALUE 2026", "PAIRS VALUE 2027",
    "PAIRS TALY% VS 2025", "PAIRS TALY% VS 2026",
    "PAIRS CONTRIBUTION % 2025", "PAIRS CONTRIBUTION % 2026", "PAIRS CONTRIBUTION % 2027",
    SEP,
    "INITIAL STOCK QTY VALUE 2025 H1", "FINAL STOCK QTY VALUE 2025 H1", "% FINAL STOCK VS INITIAL 25H1",
    "INITIAL STOCK QTY VALUE 2026 H1", "FINAL STOCK QTY VALUE 2026 H1", "% FINAL STOCK VS INITIAL 26H1",
    "INITIAL STOCK QTY VALUE 2027 H1", "FINAL STOCK QTY VALUE 2027 H1", "% FINAL STOCK VS INITIAL 27H1",
    SEP,
    "OTB QTY 2026 H1", "OTB COST 2026 H1", "SOR FINAL 2026 H2",
    "OTB QTY 2027 H1", "OTB COST 2027 H1", "SOR TARGET 2027H1",
    SEP,
    "STOCK TURN 2025", "STOCK TURN 2026", "STOCK TURN 2027",
]

# ==========================================================
# DISPLAY-ONLY HEADER ABBREVIATIONS (UI ONLY - NEVER USED FOR
# CSV/EXCEL WRITING OR CALCULATIONS)
# ==========================================================
DISPLAY_HEADER_LABELS: dict[str, str] = {
    "TURNOVER VALUE 2025": "TO 2025",
    "TURNOVER VALUE 2026": "TO 2026",
    "TURNOVER VALUE 2027": "TO 2027",
    "TO TALY% VS 2025": "TO TALY 25",
    "TO TALY% VS 2026": "TO TALY 26",
    "TO CONTRIBUTION % 2025": "TO CONT 25",
    "TO CONTRIBUTION % 2026": "TO CONT 26",
    "TO CONTRIBUTION % 2027": "TO CONT 27",
    "INTAKE% 2025": "INT % 25",
    "INTAKE% 2026": "INT % 26",
    "INTAKE% 2027": "INT % 27",
    "INTAKE TALY% VS 2025": "INT TALY 25",
    "INTAKE TALY% VS 2026": "INT TALY 26",
    "MKD% 2025": "MKD % 25",
    "MKD% 2026": "MKD % 26",
    "MKD% 2027": "MKD % 27",
    "MRG% 2025": "MRG % 25",
    "MRG% 2026": "MRG % 26",
    "MRG% 2027": "MRG % 27",
    "RRP 2025": "RRP 25",
    "RRP 2026": "RRP 26",
    "RRP 2027": "RRP 27",
    "RRP TALY% VS 2025": "RRP TALY 25",
    "RRP TALY% VS 2026": "RRP TALY 26",
    "RRP VAT 2025": "RRP VAT 25",
    "RRP VAT 2026": "RRP VAT 26",
    "RRP VAT 2027": "RRP VAT 27",
    "RRP VAT TALY% VS 2025": "RRP VAT TALY 25",
    "RRP VAT TALY% VS 2026": "RRP VAT TALY 26",
    "ASP 2025": "ASP 25",
    "ASP 2026": "ASP 26",
    "ASP 2027": "ASP 27",
    "ASP TALY% VS 2025": "ASP TALY 25",
    "ASP TALY% VS 2026": "ASP TALY 26",
    "MKD VALUE 2025": "MKD VAL 25",
    "MKD VALUE 2026": "MKD VAL 26",
    "MKD VALUE 2027": "MKD VAL 27",
    "MRG VALUE 2025": "MRG VAL 25",
    "MRG VALUE 2026": "MRG VAL 26",
    "MRG VALUE 2027": "MRG VAL 27",
    "COGS VALUE 2025": "COGS 25",
    "COGS VALUE 2026": "COGS 26",
    "COGS VALUE 2027": "COGS 27",
    "GROSS VALUE 2025": "GROSS 25",
    "GROSS VALUE 2026": "GROSS 26",
    "GROSS VALUE 2027": "GROSS 27",
    "PAIRS VALUE 2025": "VOL 25",
    "PAIRS VALUE 2026": "VOL 26",
    "PAIRS VALUE 2027": "VOL 27",
    "PAIRS TALY% VS 2025": "VOL TALY 25",
    "PAIRS TALY% VS 2026": "VOL TALY 26",
    "PAIRS CONTRIBUTION % 2025": "VOL CONT 25",
    "PAIRS CONTRIBUTION % 2026": "VOL CONT 26",
    "PAIRS CONTRIBUTION % 2027": "VOL CONT 27",
    "INITIAL STOCK QTY VALUE 2025 H1": "INIT STK 25",
    "FINAL STOCK QTY VALUE 2025 H1": "FIN STK 25",
    "% FINAL STOCK VS INITIAL 25H1": "STK VAR 25",
    "INITIAL STOCK QTY VALUE 2026 H1": "INIT STK 26",
    "FINAL STOCK QTY VALUE 2026 H1": "FIN STK 26",
    "% FINAL STOCK VS INITIAL 26H1": "STK VAR 26",
    "INITIAL STOCK QTY VALUE 2027 H1": "INIT STK 27",
    "FINAL STOCK QTY VALUE 2027 H1": "FIN STK 27",
    "% FINAL STOCK VS INITIAL 27H1": "STK VAR 27",
    "OTB QTY 2026 H1": "OTB QTY 26",
    "OTB COST 2026 H1": "OTB COST 26",
    "SOR FINAL 2026 H2": "SOR 26H2",
    "OTB QTY 2027 H1": "OTB QTY 27",
    "OTB COST 2027 H1": "OTB COST 27",
    "SOR TARGET 2027H1": "SOR 27H1",
    "STOCK TURN 2025": "STK TURN 25",
    "STOCK TURN 2026": "STK TURN 26",
    "STOCK TURN 2027": "STK TURN 27",
}

def display_header_label(column: str) -> str:
    return DISPLAY_HEADER_LABELS.get(column, column)

RATIO_TALY_CF_COLUMNS = {
    "TO TALY% VS 2026",
    "RRP TALY% VS 2026",
    "RRP VAT TALY% VS 2026",
    "ASP TALY% VS 2026",
    "PAIRS TALY% VS 2026",
}
INTAKE_TALY_CF_COLUMNS = {"INTAKE TALY% VS 2026"}

# ==========================================================
# FACTORS
# ==========================================================
FACTOR_OPTIONS = {
    "*1": 1,
    "*10": 10,
    "*100": 100,
    "*1000": 1000,
    "*10000": 10000,
    "*100000": 100000,
    "*1000000": 1000000,
    "/1": 1,
    "/10": 0.1,
    "/100": 0.01,
    "/1000": 0.001,
    "/10000": 0.0001,
    "/100000": 0.00001,
    "/1000000": 0.000001,
}
FACTOR_OPTIONS_LIST = list(FACTOR_OPTIONS.keys())

# ==========================================================
# FX TABLE
# ==========================================================
RATES_RAW = {
    "EUR_GROUP": 0.878, "CZK": 22.027, "CHF": 0.827, "GBP": 0.78, "AUD": 1.535,
    "BDT": 124.652, "BOB": 6.91, "BWP": 13.587, "CAD": 1.386, "CLP": 939.085,
    "CNY": 7.242, "COP": 4312.63, "USD": 1.0, "INR": 86.038, "IDR": 16430.45,
    "KES": 136.217, "MWK": 1734.58, "MYR": 4.306, "MXN": 17.496, "NZD": 1.831,
    "PKR": 287.545, "PHP": 55.663, "PEN": 3.72, "SGD": 1.316, "ZAR": 18.306,
    "LKR": 299.273, "TZS": 2681.25, "THB": 33.208, "UGX": 3718.53, "VND": 25967.95,
    "ZMW": 27.833,
}
COUNTRY_ALIASES = {
    "ITALY": "EUR_GROUP", "ITALIA": "EUR_GROUP", "ITA": "EUR_GROUP",
    "FRANCE": "EUR_GROUP", "FRANCIA": "EUR_GROUP", "FRA": "EUR_GROUP",
    "SPAIN": "EUR_GROUP", "ESPANA": "EUR_GROUP", "ESP": "EUR_GROUP",
    "SLOVAKIA": "EUR_GROUP", "ESLOVAQUIA": "EUR_GROUP", "SVK": "EUR_GROUP",
    "NETHERLANDS": "EUR_GROUP", "HOLANDA": "EUR_GROUP", "NLD": "EUR_GROUP",
    "TURKIYE": "EUR_GROUP", "TURKEY": "EUR_GROUP", "TURQUIA": "EUR_GROUP", "TUR": "EUR_GROUP",
    "CZECH REPUBLIC": "CZK", "REPUBLICA CHECA": "CZK", "CZE": "CZK", "CZECH": "CZK",
    "SWITZERLAND": "CHF", "SUIZA": "CHF", "CHE": "CHF",
    "UNITED KINGDOM": "GBP", "REINO UNIDO": "GBP", "UK": "GBP", "GBR": "GBP",
    "AUSTRALIA": "AUD", "AUS": "AUD",
    "BANGLADESH": "BDT", "BGD": "BDT",
    "BOLIVIA": "BOB", "BOL": "BOB",
    "BOTSWANA": "BWP", "BWA": "BWP",
    "CANADA": "CAD", "CAN": "CAD",
    "CHILE": "CLP", "CHI": "CLP",
    "CHINA": "CNY", "CHN": "CNY",
    "COLOMBIA": "COP", "COL": "COP",
    "ECUADOR": "USD", "ECU": "USD",
    "INDIA": "INR", "IND": "INR",
    "INDONESIA": "IDR", "IDN": "IDR",
    "KENYA": "KES", "KEN": "KES",
    "MALAWI": "MWK", "MWI": "MWK",
    "MALAYSIA": "MYR", "MYS": "MYR",
    "MEXICO": "MXN", "MEX": "MXN",
    "NEW ZEALAND": "NZD", "NZL": "NZD",
    "PAKISTAN": "PKR", "PAK": "PKR",
    "PHILIPPINES": "PHP", "FILIPINAS": "PHP", "PHL": "PHP",
    "PERU": "PEN", "PER": "PEN",
    "SINGAPORE": "SGD", "SINGAPUR": "SGD", "SGP": "SGD",
    "SOUTH AFRICA": "ZAR", "SUDAFRICA": "ZAR", "ZAF": "ZAR",
    "SRI LANKA": "LKR", "LKA": "LKR",
    "TANZANIA": "TZS", "TZA": "TZS",
    "THAILAND": "THB", "TAILANDIA": "THB", "THA": "THB",
    "UGANDA": "UGX", "UGA": "UGX",
    "VIETNAM": "VND", "VNM": "VND",
    "ZAMBIA": "ZMW", "ZMB": "ZMW",
    "ZIMBABWE": "USD", "ZWE": "USD",
}
ISO2_ALIASES = {
    "IT": "EUR_GROUP", "FR": "EUR_GROUP", "ES": "EUR_GROUP", "SK": "EUR_GROUP",
    "NL": "EUR_GROUP", "TR": "EUR_GROUP", "CZ": "CZK", "CH": "CHF", "GB": "GBP",
    "AU": "AUD", "BD": "BDT", "BO": "BOB", "BW": "BWP", "CA": "CAD", "CL": "CLP",
    "CN": "CNY", "CO": "COP", "EC": "USD", "IN": "INR", "ID": "IDR", "KE": "KES",
    "MW": "MWK", "MY": "MYR", "MX": "MXN", "NZ": "NZD", "PK": "PKR", "PH": "PHP",
    "PE": "PEN", "SG": "SGD", "ZA": "ZAR", "LK": "LKR", "TZ": "TZS", "TH": "THB",
    "UG": "UGX", "VN": "VND", "ZM": "ZMW", "ZW": "USD",
}

# ==========================================================
# USERS
# ==========================================================
USERS = {
    "ADMIN": {"password": "Global123", "country": "ALL", "role": "admin"},
    "BANGLADESH": {"password": "Password123", "country": "BANGLADESH", "role": "country"},
    "BOLIVIA": {"password": "Password123", "country": "BOLIVIA", "role": "country"},
    "CHILE": {"password": "Password123", "country": "CHILE", "role": "country"},
    "COLOMBIA": {"password": "Password123", "country": "COLOMBIA", "role": "country"},
    "CZECH": {"password": "Password123", "country": "CZECH", "role": "country"},
    "ECUADOR": {"password": "Password123", "country": "ECUADOR", "role": "country"},
    "INDIA": {"password": "Password123", "country": "INDIA", "role": "country"},
    "INDONESIA": {"password": "Password123", "country": "INDONESIA", "role": "country"},
    "ITALIA": {"password": "Password123", "country": "ITALIA", "role": "country"},
    "KENYA": {"password": "Password123", "country": "KENYA", "role": "country"},
    "MALAYSIA": {"password": "Password123", "country": "MALAYSIA", "role": "country"},
    "PAKISTAN": {"password": "Password123", "country": "PAKISTAN", "role": "country"},
    "PERU": {"password": "Password123", "country": "PERU", "role": "country"},
    "SINGAPUR": {"password": "Password123", "country": "SINGAPUR", "role": "country"},
    "SLOVAKIA": {"password": "Password123", "country": "SLOVAKIA", "role": "country"},
    "SPAIN": {"password": "Password123", "country": "SPAIN", "role": "country"},
    "SWITZERLAND": {"password": "Password123", "country": "SWITZERLAND", "role": "country"},
    "THAILAND": {"password": "Password123", "country": "THAILAND", "role": "country"},
}
COUNTRY_USERS = [k for k in USERS if k != "ADMIN"]
MENU_OPTIONS = COUNTRY_USERS + ["ADMIN"]
COUNTRY_CURRENCIES = {
    "BANGLADESH": ["BDT", "USD"], "BOLIVIA": ["BOB", "USD"], "CHILE": ["CLP", "USD"],
    "COLOMBIA": ["COP", "USD"], "CZECH": ["CZK", "USD"], "ECUADOR": ["USD"],
    "INDIA": ["INR", "USD"], "INDONESIA": ["IDR", "USD"], "ITALIA": ["EUR", "USD"],
    "KENYA": ["KES", "USD"], "MALAYSIA": ["MYR", "USD"], "PAKISTAN": ["PKR", "USD"],
    "PERU": ["PEN", "USD"], "SINGAPUR": ["SGD", "USD"], "SLOVAKIA": ["EUR", "USD"],
    "SPAIN": ["EUR", "USD"], "SWITZERLAND": ["CHF", "USD"], "THAILAND": ["THB", "USD"],
}
DEFAULT_VAT_RATES = {
    "BANGLADESH": 15.0, "BOLIVIA": 14.95, "CHILE": 19.0, "COLOMBIA": 19.0,
    "CZECH": 21.0, "ECUADOR": 15.0, "INDIA": 18.0, "INDONESIA": 12.0,
    "ITALIA": 22.0, "KENYA": 16.0, "MALAYSIA": 10.0, "PAKISTAN": 18.0,
    "PERU": 18.0, "SINGAPUR": 9.0, "SLOVAKIA": 23.0, "SPAIN": 21.0,
    "SWITZERLAND": 8.1, "THAILAND": 7.0,
}
ALL_CURRENCY_OPTIONS = sorted(dict.fromkeys(COUNTRY_CURRENCIES[c][0] for c in COUNTRY_CURRENCIES)) + ["USD"]

# ==========================================================
# PAGE CONFIG + CSS
# ==========================================================
st.set_page_config(
    page_title="Global Category Plan",
    page_icon="📊",
    layout="wide",
)
st.markdown(
    """<style>
:root{--bata-red:#ed1b2f;--bata-dark:#252525;--border:#e5e8ee;--blue:#2667d9}
html,body,[data-testid="stAppViewContainer"],[data-testid="stMain"]{background:#fff!important}
[data-testid="stHeader"]{background:transparent!important}
[data-testid="stToolbar"],[data-testid="stDecoration"],[aria-label="Deploy"],[aria-label="Main menu"],[aria-label="More options"]{display:none!important}
.block-container{max-width:100%!important;padding:.1rem 1rem 1rem!important}
.app-topbar{background:#fff;border-bottom:1px solid var(--border);display:flex;align-items:center;margin:-.1rem -1rem 2px;padding:2px 24px;box-shadow:0 1px 4px rgba(0,0,0,.05)}
.app-topbar{margin-bottom:0 !important}
.brand-compact{font-size:10px;font-weight:700;color:var(--bata-dark);margin:0}
.top-header-divider{
    width:100%;
    border-bottom:1px solid #D1D5DB;
    margin:3px 0 16px 0;
}

/* =========================================================
   NAVIGATION BUTTONS
   ========================================================= */
.nav-button-slot button{
    background-color:#E8F0FE!important;
    border:1px solid #B9D3F7!important;
    border-radius:6px!important;
    color:#1449A6!important;
    font-size:14px!important;
    font-weight:600!important;
    min-height:36px!important;
}
.nav-button-slot.active button{
    background-color:#D6E4FB!important;
    color:var(--blue)!important;
}
/* Separator line before Log Out */
.logout-separator{
    border-left:1px solid #D1D5DB !important;
    height:32px !important;
    margin:0 8px !important;
}

.control-label-spacer{
    height:28px;
}

.dashboard-main-title{font-size:24px;font-weight:800;color:var(--bata-dark);margin:2px 0 6px 0;padding:0}
.dashboard-section-title{font-size:16px;font-weight:700;color:var(--bata-dark);margin:.2rem 0 .1rem}
.dashboard-section-subtitle{color:#657083;font-size:13px;margin-bottom:.35rem}
[data-testid="stMetric"]{background:#fff;border:1px solid var(--border);border-radius:10px;padding:8px 10px;min-height:85px}
[data-testid="stMetricLabel"]{font-size:12px!important;font-weight:600!important}
[data-testid="stMetricValue"]{font-size:22px!important}
[data-testid="stMetricDelta"]{font-size:12px!important;font-weight:600}
div[data-testid="stMultiSelect"] span{font-size:12px}
.compact-row label{font-size:12px!important;margin-bottom:.10rem!important}
.compact-row [data-testid="stMultiSelect"]{margin-bottom:0!important}
div[data-testid="stForm"]{border:none!important;padding:0!important}

/* =========================================================
   GLOBAL MULTISELECT SELECTED TAGS — LIGHT PASTEL PINK
   ========================================================= */

div[data-testid="stMultiSelect"] [data-baseweb="tag"],
div[data-testid="stMultiSelect"] div[data-baseweb="tag"],
div[data-testid="stMultiSelect"] span[data-baseweb="tag"],
div[data-baseweb="select"] [data-baseweb="tag"],
div[data-baseweb="tag"],
span[data-baseweb="tag"],
[data-baseweb="tag"] {
    background: #F8CACA !important;
    background-color: #F8CACA !important;
    border: 1px solid #F8CACA !important;
    border-color: #F8CACA !important;
    color: #1B1B1B !important;
    box-shadow: none !important;
    opacity: 1 !important;
}

div[data-testid="stMultiSelect"] [data-baseweb="tag"] *,
div[data-baseweb="tag"] *,
span[data-baseweb="tag"] *,
[data-baseweb="tag"] * {
    color: #1B1B1B !important;
    fill: #1B1B1B !important;
    stroke: #1B1B1B !important;
    -webkit-text-fill-color: #1B1B1B !important;
}

div[data-testid="stMultiSelect"] [data-baseweb="tag"] svg,
div[data-testid="stMultiSelect"] [data-baseweb="tag"] svg *,
div[data-baseweb="tag"] svg,
div[data-baseweb="tag"] svg *,
span[data-baseweb="tag"] svg,
span[data-baseweb="tag"] svg * {
    color: #1B1B1B !important;
    fill: #1B1B1B !important;
    stroke: #1B1B1B !important;
}

/* Selected / highlighted item inside open dropdown menu */
div[role="listbox"] [aria-selected="true"],
ul[role="listbox"] [aria-selected="true"],
li[aria-selected="true"] {
    background: #FDE2E2 !important;
    background-color: #FDE2E2 !important;
    color: #1B1B1B !important;
}

/* Filter multiselect height matching segmented control */
div[data-testid="stMultiSelect"] > div[data-baseweb="select"] {
    height: 36px !important;
    min-height: 36px !important;
}

/* =========================================================
   FOOTER
   ========================================================= */
.app-footer {
    margin-top: 28px;
    padding: 14px 0 4px 0;
    border-top: 1px solid #E5E7EB;
    text-align: center;
    color: #6B7280;
    font-size: 12px;
}
.app-footer a {
    color: #2667D9;
    text-decoration: none;
}
.app-footer a:hover {
    text-decoration: underline;
}
</style>""",
    unsafe_allow_html=True,
)

# ==========================================================
# TEXT / NUMBER UTILITIES
# ==========================================================
def clean_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    return "" if text.lower() in {"none", "nan", "null", ""} else text

def parse_number(value) -> float | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if not text or text.startswith("="):
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None

def safe_divide(a, b) -> float:
    if a is None or b is None:
        return 0.0
    try:
        if float(b) == 0:
            return 0.0
        return float(a) / float(b)
    except (TypeError, ValueError):
        return 0.0

def series_div(numerator, denominator):
    num = pd.to_numeric(numerator, errors="coerce")

    if isinstance(denominator, pd.Series):
        den = pd.to_numeric(denominator, errors="coerce").mask(
            pd.to_numeric(denominator, errors="coerce") == 0
        )
        return num / den

    den = pd.to_numeric(denominator, errors="coerce")

    try:
        if pd.isna(den) or float(den) == 0:
            if isinstance(num, pd.Series):
                return pd.Series(pd.NA, index=num.index, dtype="Float64")
            return pd.NA
    except (TypeError, ValueError):
        if isinstance(num, pd.Series):
            return pd.Series(pd.NA, index=num.index, dtype="Float64")
        return pd.NA

    return num / den

def header_token(value) -> str:
    return re.sub(r"\s+", " ", clean_text(value).upper())

def is_total_text(value) -> bool:
    text = header_token(value)
    if not text:
        return False
    compact = re.sub(r"[-_/]+", " ", text)
    compact = re.sub(r"\s+", " ", compact).strip()
    if compact in {"TOTAL", "TOT", "SUBTOTAL", "SUB TOTAL"}:
        return True
    if compact.startswith(("TOTAL ", "TOT ", "SUBTOTAL ", "SUB TOTAL ")):
        return True
    return False

def normalize_text(value: str) -> str:
    text = str(value).strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", text)

def get_fx_rate(country: str) -> float:
    key = normalize_text(country)
    if key in COUNTRY_ALIASES:
        return RATES_RAW[COUNTRY_ALIASES[key]]
    if key in ISO2_ALIASES:
        return RATES_RAW[ISO2_ALIASES[key]]
    for token in key.split(" "):
        if token in COUNTRY_ALIASES:
            return RATES_RAW[COUNTRY_ALIASES[token]]
        if len(token) == 2 and token in ISO2_ALIASES:
            return RATES_RAW[ISO2_ALIASES[token]]
    raise ValueError(f"No FX rate configured for country: {country}")

def format_integer(value) -> str:
    if value is None or pd.isna(value):
        return "0"
    try:
        return f"{float(value):,.0f}".replace(",", ".")
    except (TypeError, ValueError):
        return clean_text(value)

def format_decimal(value, decimals: int = 2) -> str:
    if value is None or pd.isna(value):
        return f"0.{'0' * decimals}"
    try:
        parts = f"{float(value):,.{decimals}f}".split(".")
        parts[0] = parts[0].replace(",", ".")
        return ",".join(parts)
    except (TypeError, ValueError):
        return clean_text(value)

def format_percent(value, decimals: int = 1) -> str:
    if value is None or pd.isna(value):
        return f"0,{('0' * decimals)}%"
    try:
        parts = f"{float(value) * 100:,.{decimals}f}".split(".")
        parts[0] = parts[0].replace(",", ".")
        return ",".join(parts) + "%"
    except (TypeError, ValueError):
        return clean_text(value)

def calc_growth(new_value, old_value) -> float:
    nv = pd.to_numeric(new_value, errors="coerce")
    ov = pd.to_numeric(old_value, errors="coerce")
    try:
        if pd.isna(nv) or pd.isna(ov):
            return 0.0
        if float(ov) == 0:
            return 0.0
        return float(nv) / float(ov) - 1.0
    except (TypeError, ValueError):
        return 0.0

def format_delta_kpi(growth_val, ref_year: str = "2026") -> str | None:
    if growth_val is None or pd.isna(growth_val):
        return None
    try:
        val_pct = float(growth_val) * 100.0
        sign = "+" if val_pct > 0 else ""
        formatted = f"{sign}{val_pct:.1f}%".replace(".", ",")
        return f"{formatted} vs {ref_year}"
    except (TypeError, ValueError):
        return None

def format_pp_delta(delta_val, ref_year: str = "2026") -> str | None:
    if delta_val is None or pd.isna(delta_val):
        return None
    try:
        points = float(delta_val) * 100.0
        sign = "+" if points > 0 else ""
        formatted = f"{sign}{points:.1f}".replace(".", ",")
        return f"{formatted} pp vs {ref_year}"
    except (TypeError, ValueError):
        return None

# ==========================================================
# COLUMN METADATA UTILITIES
# ==========================================================
def normalize_column_name(name: str) -> str:
    return str(name).replace("\ufeff", "").strip().upper()

def find_column(data, expected_name: str) -> str | None:
    expected = normalize_column_name(expected_name)
    for column in data.columns:
        if normalize_column_name(column) == expected:
            return column
    return None

def is_separator_column(column: str) -> bool:
    name = str(column)
    if name != "" and set(name) <= {"\u200b"}:
        return True
    return bool(re.fullmatch(r"VACIA\d*", normalize_column_name(column)))

def is_text_column(column: str) -> bool:
    return normalize_column_name(column) in {"KEY", "BU", "GENDER", "OCCASION", "CATEGORY", "COUNTRY"}

def is_omitted_column(column: str) -> bool:
    return normalize_column_name(column) in {normalize_column_name(c) for c in OMIT_COLUMNS}

def is_percentage_column(column: str) -> bool:
    if is_omitted_column(column):
        return False
    normalized = normalize_column_name(column)
    if "STOCK TURN" in normalized:
        return False
    return any(
        tok in normalized
        for tok in ["%", "TALY", "CONTRIBUTION", "FINAL STOCK VS INITIAL", "SOR"]
    )

def is_decimal_metric_column(column: str) -> bool:
    normalized = normalize_column_name(column)
    return normalized.startswith("RRP") or normalized.startswith("ASP") or "STOCK TURN" in normalized

def is_count_or_money_column(column: str) -> bool:
    normalized = normalize_column_name(column)
    return any(
        token in normalized
        for token in [
            "VALUE", "TURNOVER", "PAIRS", "QTY", "COST", "RRP",
            "ASP", "STOCK TURN", "OTB", "SOR", "MKD", "MRG", "COGS", "GROSS",
        ]
    )

def make_sep_name(index: int) -> str:
    return "\u200b" * (index + 1)

# ==========================================================
# METADATA
# ==========================================================
def load_upload_metadata() -> dict:
    if not UPLOAD_META_PATH.exists():
        return {}
    try:
        return json.loads(UPLOAD_META_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}

def save_upload_metadata(meta: dict) -> None:
    UPLOAD_META_PATH.write_text(json.dumps(meta, indent=2), encoding="utf-8")

def vat_rate_for_data(data: pd.DataFrame) -> float:
    meta = load_upload_metadata()
    if data is None or data.empty or "COUNTRY" not in data.columns:
        rates = [float(v.get("vat_rate", 0.0) or 0.0) for v in meta.values()]
        return rates[0] if rates and all(abs(r - rates[0]) < 1e-12 for r in rates) else 0.0
    countries = [str(c).upper() for c in data["COUNTRY"].dropna().unique().tolist() if str(c).strip()]
    rates = []
    weights = []
    turnover_col = find_column(data, "TURNOVER VALUE 2027")
    for country in countries:
        entry = meta.get(country, {})
        rate = float(entry.get("vat_rate", 0.0) or 0.0)
        rates.append(rate)
        if turnover_col:
            weight = float(pd.to_numeric(data.loc[data["COUNTRY"].astype(str).str.upper() == country, turnover_col], errors="coerce").fillna(0).sum())
        else:
            weight = 1.0
        weights.append(max(weight, 0.0))
    if not rates:
        return 0.0
    if all(abs(r - rates[0]) < 1e-12 for r in rates):
        return rates[0]
    total_w = sum(weights)
    if total_w <= 0:
        return sum(rates) / len(rates)
    return sum(r * w for r, w in zip(rates, weights)) / total_w

# ==========================================================
# EXCEL SOURCE READING
# ==========================================================
def workbook_sheet_names(file_bytes: bytes) -> list[str]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=False)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()

def find_bu_column(ws) -> int:
    for col_idx in range(1, 41):
        if header_token(ws.cell(row=7, column=col_idx).value) == "BU":
            return col_idx
    raise ValueError(
        "Could not locate the 'BU' header in row 7. "
        "Please check that row 7 contains BU, GENDER, OCCASION and CATEGORY in that order."
    )

def validate_source_headers(ws) -> int:
    bu_col = find_bu_column(ws)
    b7 = header_token(ws.cell(row=7, column=bu_col).value)
    c7 = header_token(ws.cell(row=7, column=bu_col + 1).value)
    d7 = header_token(ws.cell(row=7, column=bu_col + 2).value)
    e7 = header_token(ws.cell(row=7, column=bu_col + 3).value)
    if b7 != "BU" or c7 != "GENDER" or d7 != "OCCASION" or e7 != "CATEGORY":
        raise ValueError(
            "The selected worksheet is not a valid category plan template. "
            "Expected BU, GENDER, OCCASION, CATEGORY in that order starting at the BU column. "
            f"Found BU={b7 or '(empty)'}, GENDER={c7 or '(empty)'}, "
            f"OCCASION={d7 or '(empty)'}, CATEGORY={e7 or '(empty)'}"
        )
    return bu_col

def row_is_fully_blank(ws, excel_row: int, col_country: int) -> bool:
    for col_idx in range(1, col_country + 1):
        value = ws.cell(row=excel_row, column=col_idx).value
        if value not in (None, ""):
            return False
    return True

def read_uploaded_workbook(file_bytes: bytes, sheet_name: str) -> list[dict]:
    names = workbook_sheet_names(file_bytes)
    if sheet_name not in names:
        raise ValueError(f"Worksheet '{sheet_name}' was not found in the uploaded file.")
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=False)
    try:
        ws = wb[sheet_name]
        bu_col = validate_source_headers(ws)
        delta = bu_col - COL_BU
        col_bu = bu_col
        col_gender = COL_GENDER + delta
        col_occasion = COL_OCCASION + delta
        col_category = COL_CATEGORY + delta
        col_country = COL_COUNTRY + delta
        source_base_map = {
            get_column_letter(column_index_from_string(letter) + delta): header
            for letter, header in SOURCE_BASE_MAP.items()
        }
        last_bu = ""
        last_gender = ""
        last_occasion = ""
        rows = []
        max_row = ws.max_row or SOURCE_START_ROW
        for excel_row in range(SOURCE_START_ROW, max_row + 1):
            if row_is_fully_blank(ws, excel_row, col_country):
                continue
            raw_bu = ws.cell(row=excel_row, column=col_bu).value
            raw_gender = ws.cell(row=excel_row, column=col_gender).value
            raw_occasion = ws.cell(row=excel_row, column=col_occasion).value
            raw_category = ws.cell(row=excel_row, column=col_category).value
            if (
                is_total_text(raw_bu)
                or is_total_text(raw_gender)
                or is_total_text(raw_occasion)
                or is_total_text(raw_category)
            ):
                continue
            if raw_bu not in (None, ""):
                last_bu = clean_text(raw_bu)
            if raw_gender not in (None, ""):
                last_gender = clean_text(raw_gender)
            if raw_occasion not in (None, ""):
                last_occasion = clean_text(raw_occasion)
            category_clean = clean_text(raw_category)
            if not category_clean:
                continue
            row_data = {
                "BU": last_bu,
                "GENDER": last_gender,
                "OCCASION": last_occasion,
                "CATEGORY": category_clean,
                "KEY": f"{last_bu}{last_gender}{last_occasion}{category_clean}",
            }
            for letter, header in source_base_map.items():
                col_idx = column_index_from_string(letter)
                row_data[header] = ws.cell(row=excel_row, column=col_idx).value
            rows.append(row_data)
    finally:
        wb.close()
    if not rows:
        raise ValueError("No valid detail rows were found in the uploaded file.")
    return rows

def validate_selected_worksheet(file_bytes: bytes, sheet_name: str) -> None:
    names = workbook_sheet_names(file_bytes)
    if sheet_name not in names:
        raise ValueError(f"Worksheet '{sheet_name}' was not found in the uploaded file.")
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=False)
    try:
        validate_source_headers(wb[sheet_name])
    finally:
        wb.close()

# ==========================================================
# NORMALIZATION
# ==========================================================
def convert_currency(value: float, input_currency: str, target: str, fx_rate: float) -> float:
    if target == "USD":
        if input_currency == "USD":
            return value
        return value / fx_rate if fx_rate else value
    if input_currency == "USD":
        return value * fx_rate if fx_rate else value
    return value

def normalize_source_rows(
    detail_rows: list[dict],
    money_factor: float,
    pairs_factor: float,
    money_columns: list[str],
    pairs_columns: list[str],
    input_currency: str,
    fx_rate: float,
    country: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    money_set = set(money_columns)  # Use selected columns, not all MONEY_COLUMNS
    pairs_set = set(pairs_columns)

    def to_base(row: dict, target_currency: str) -> dict:
        out = {header: None for header in HEADERS}
        out["KEY"] = row["KEY"]
        out["BU"] = row["BU"]
        out["GENDER"] = row["GENDER"]
        out["OCCASION"] = row["OCCASION"]
        out["CATEGORY"] = row["CATEGORY"]
        out["COUNTRY"] = country
        out["TO NEW COLLECTION"] = parse_number(row.get("TO NEW COLLECTION"))
        out["SHARE %"] = parse_number(row.get("SHARE %"))
        for header in BASE_ADDITIVE_COLUMNS:
            value = parse_number(row.get(header))
            if value is None:
                out[header] = None
                continue
            if header in money_set:
                value = value * money_factor
            if header in pairs_set:
                value = value * pairs_factor
            if header in MONEY_COLUMNS:  # Currency conversion always applies to all money columns
                value = convert_currency(value, input_currency, target_currency, fx_rate)
            out[header] = value
        return out

    df_lc = pd.DataFrame([to_base(r, "LC") for r in detail_rows], columns=HEADERS)
    df_usd = pd.DataFrame([to_base(r, "USD") for r in detail_rows], columns=HEADERS)
    return df_lc, df_usd

def aggregate_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.reindex(columns=HEADERS)
    working = df.copy()
    for column in TEXT_COLUMNS:
        if column not in working.columns:
            working[column] = ""
        working[column] = working[column].map(clean_text)
    for column in BASE_ADDITIVE_COLUMNS:
        if column not in working.columns:
            working[column] = 0.0
        working[column] = pd.to_numeric(working[column], errors="coerce").fillna(0.0)
    grouped_text = working.groupby(["COUNTRY", "KEY"], as_index=False, sort=False)[TEXT_COLUMNS].first()
    grouped_add = working.groupby(["COUNTRY", "KEY"], as_index=False, sort=False)[BASE_ADDITIVE_COLUMNS].sum()
    merged = grouped_text.merge(grouped_add, on=["COUNTRY", "KEY"], how="left")
    for column in PRESERVED_NON_ADDITIVE:
        if column in working.columns:
            first_vals = working.groupby(["COUNTRY", "KEY"], as_index=False, sort=False)[column].first()
            merged = merged.merge(first_vals, on=["COUNTRY", "KEY"], how="left")
    return merged.reindex(columns=HEADERS)

# ==========================================================
# EXCEL WRITERS
# ==========================================================
def _write_workbook_structure(ws) -> None:
    bold = Font(bold=True)
    for col_idx, type_label in enumerate(ROW2_TYPES, start=1):
        ws.cell(row=SOURCE_MAPPING_ROW, column=col_idx).value = type_label
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=SOURCE_HEADER_ROW, column=col_idx)
        cell.value = header
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[SOURCE_HEADER_ROW].height = 45
    for col_idx, header in enumerate(HEADERS, start=1):
        letter = get_column_letter(col_idx)
        width = min(max(len(str(header)) + 2, 12), 32)
        ws.column_dimensions[letter].width = width

def _write_formulas(ws, first_row: int, last_row: int, vat_rate: float) -> None:
    for r in range(first_row, last_row + 1):
        ws.cell(row=r, column=HEADER_TO_COL["TO TALY% VS 2025"]).value = f'=IFERROR(H{r}/F{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["TO TALY% VS 2026"]).value = f'=IFERROR(H{r}/G{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["TO CONTRIBUTION % 2025"]).value = f'=IFERROR(F{r}/SUM($F$4:$F$100000),"")'
        ws.cell(row=r, column=HEADER_TO_COL["TO CONTRIBUTION % 2026"]).value = f'=IFERROR(G{r}/SUM($G$4:$G$100000),"")'
        ws.cell(row=r, column=HEADER_TO_COL["TO CONTRIBUTION % 2027"]).value = f'=IFERROR(H{r}/SUM($H$4:$H$100000),"")'
        ws.cell(row=r, column=HEADER_TO_COL["INTAKE% 2025"]).value = f'=IFERROR((BI{r}-BE{r})/BI{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["INTAKE% 2026"]).value = f'=IFERROR((BJ{r}-BF{r})/BJ{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["INTAKE% 2027"]).value = f'=IFERROR((BK{r}-BG{r})/BK{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["INTAKE TALY% VS 2025"]).value = f'=IFERROR(S{r}-Q{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["INTAKE TALY% VS 2026"]).value = f'=IFERROR(S{r}-R{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["MKD% 2025"]).value = f'=IFERROR(AW{r}/BI{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["MKD% 2026"]).value = f'=IFERROR(AX{r}/BJ{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["MKD% 2027"]).value = f'=IFERROR(AY{r}/BK{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["MRG% 2025"]).value = f'=IFERROR(BA{r}/F{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["MRG% 2026"]).value = f'=IFERROR(BB{r}/G{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["MRG% 2027"]).value = f'=IFERROR(BC{r}/H{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["RRP 2025"]).value = f'=IFERROR(BI{r}/BM{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["RRP 2026"]).value = f'=IFERROR(BJ{r}/BN{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["RRP 2027"]).value = f'=IFERROR(BK{r}/BO{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["RRP TALY% VS 2025"]).value = f'=IFERROR(AG{r}/AE{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["RRP TALY% VS 2026"]).value = f'=IFERROR(AG{r}/AF{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["RRP VAT 2025"]).value = f'=IFERROR(AE{r}*(1+{vat_rate}),"")'
        ws.cell(row=r, column=HEADER_TO_COL["RRP VAT 2026"]).value = f'=IFERROR(AF{r}*(1+{vat_rate}),"")'
        ws.cell(row=r, column=HEADER_TO_COL["RRP VAT 2027"]).value = f'=IFERROR(AG{r}*(1+{vat_rate}),"")'
        ws.cell(row=r, column=HEADER_TO_COL["RRP VAT TALY% VS 2025"]).value = f'=IFERROR(AM{r}/AK{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["RRP VAT TALY% VS 2026"]).value = f'=IFERROR(AM{r}/AL{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["ASP 2025"]).value = f'=IFERROR(F{r}/BM{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["ASP 2026"]).value = f'=IFERROR(G{r}/BN{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["ASP 2027"]).value = f'=IFERROR(H{r}/BO{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["ASP TALY% VS 2025"]).value = f'=IFERROR(AS{r}/AQ{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["ASP TALY% VS 2026"]).value = f'=IFERROR(AS{r}/AR{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["PAIRS TALY% VS 2025"]).value = f'=IFERROR(BO{r}/BM{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["PAIRS TALY% VS 2026"]).value = f'=IFERROR(BO{r}/BN{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["PAIRS CONTRIBUTION % 2025"]).value = f'=IFERROR(BM{r}/SUM($BM$4:$BM$100000),"")'
        ws.cell(row=r, column=HEADER_TO_COL["PAIRS CONTRIBUTION % 2026"]).value = f'=IFERROR(BN{r}/SUM($BN$4:$BN$100000),"")'
        ws.cell(row=r, column=HEADER_TO_COL["PAIRS CONTRIBUTION % 2027"]).value = f'=IFERROR(BO{r}/SUM($BO$4:$BO$100000),"")'
        ws.cell(row=r, column=HEADER_TO_COL["% FINAL STOCK VS INITIAL 25H1"]).value = f'=IFERROR(BW{r}/BV{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["% FINAL STOCK VS INITIAL 26H1"]).value = f'=IFERROR(BZ{r}/BY{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["% FINAL STOCK VS INITIAL 27H1"]).value = f'=IFERROR(CC{r}/CB{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["SOR FINAL 2026 H2"]).value = f'=IFERROR(BN{r}/(BN{r}+BZ{r}),0)'
        ws.cell(row=r, column=HEADER_TO_COL["SOR TARGET 2027H1"]).value = f'=IFERROR(CC{r}/CI{r},"")'
        ws.cell(row=r, column=HEADER_TO_COL["STOCK TURN 2025"]).value = f'=IFERROR(BM{r}/AVERAGE(BV{r}:BW{r}),"")'
        ws.cell(row=r, column=HEADER_TO_COL["STOCK TURN 2026"]).value = f'=IFERROR(BN{r}/AVERAGE(BY{r}:BZ{r}),"")'
        ws.cell(row=r, column=HEADER_TO_COL["STOCK TURN 2027"]).value = f'=IFERROR(BO{r}/AVERAGE(CB{r}:CC{r}),"")'

def _write_data_row(ws, df: pd.DataFrame, first_row: int) -> int:
    write_headers = TEXT_COLUMNS + BASE_ADDITIVE_COLUMNS + PRESERVED_NON_ADDITIVE
    row = first_row
    for _, data_row in df.iterrows():
        for header in write_headers:
            col_idx = HEADER_TO_COL.get(header)
            if not col_idx:
                continue
            value = data_row.get(header)
            cell = ws.cell(row=row, column=col_idx)
            if value is None or (isinstance(value, float) and pd.isna(value)):
                cell.value = None
            else:
                cell.value = value
        row += 1
    return row

def _write_metadata_sheet(wb, metadata_rows: list[dict]) -> None:
    if METADATA_SHEET_NAME in wb.sheetnames:
        del wb[METADATA_SHEET_NAME]
    ws = wb.create_sheet(METADATA_SHEET_NAME)
    headers = [
        "COUNTRY", "UPLOAD_DATE", "SOURCE_FILE", "CURRENCY", "FX_RATE",
        "MONEY_FACTOR", "PAIRS_FACTOR", "VAT_RATE", "ROW_COUNT",
    ]
    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = bold
    for row_idx, meta in enumerate(metadata_rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx).value = meta.get(header)
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

def save_normalized_excel(df: pd.DataFrame, target: Path, metadata_rows: list[dict], vat_rate: float) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = GLOBAL_SHEET_NAME
    _write_workbook_structure(ws)
    last_row = _write_data_row(ws, df, SOURCE_DATA_START_ROW) - 1
    if last_row >= SOURCE_DATA_START_ROW:
        _write_formulas(ws, SOURCE_DATA_START_ROW, last_row, vat_rate)
    _write_metadata_sheet(wb, metadata_rows)
    wb.save(target)
    wb.close()
    return target

def write_global_csv(df: pd.DataFrame, path: Path) -> None:
    export = df.reindex(columns=HEADERS).copy()
    export.to_csv(path, sep=";", decimal=",", index=False, encoding="utf-8-sig")

def read_global_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=HEADERS)
    data = pd.read_csv(
        path,
        sep=";",
        decimal=",",
        encoding="utf-8-sig",
        keep_default_na=False,
    )
    data.columns = [str(c).replace("\ufeff", "").strip() for c in data.columns]
    return data.reindex(columns=HEADERS)

def upsert_country_frames(existing: pd.DataFrame, new_df: pd.DataFrame, country: str) -> pd.DataFrame:
    if existing is None or existing.empty:
        combined = new_df.copy()
    else:
        working = existing.copy()
        if "COUNTRY" in working.columns:
            working = working[working["COUNTRY"].astype(str).str.upper() != country.upper()]
        combined = pd.concat([working, new_df], ignore_index=True)
    return combined.reindex(columns=HEADERS)

def _backup_file(path: Path) -> Path | None:
    if not path.exists():
        return None
    backup = path.with_name(path.name + ".bak")
    shutil.copy2(path, backup)
    return backup

def _restore_outputs(backups: dict[Path, Path | None], created: list[Path]) -> None:
    for original, backup in backups.items():
        if backup and backup.exists():
            shutil.copy2(backup, original)
        elif original.exists() and original in created:
            original.unlink()
    for path in created:
        if path not in backups and path.exists():
            path.unlink()

def _cleanup_backups(backups: dict[Path, Path | None]) -> None:
    for backup in backups.values():
        if backup and backup.exists():
            backup.unlink()

def collect_metadata_rows_from_meta(meta: dict) -> list[dict]:
    return [dict(entry, COUNTRY=entry.get("COUNTRY", country)) for country, entry in meta.items()]

def collect_metadata_rows(current_country: str, current_meta: dict) -> list[dict]:
    stored = load_upload_metadata()
    stored[current_country.upper()] = current_meta
    return collect_metadata_rows_from_meta(stored)

def remove_country_data(country: str) -> None:
    country_upper = country.upper()
    existing_usd = read_global_csv(GLOBAL_DATA_CSV)
    if not existing_usd.empty and "COUNTRY" in existing_usd.columns:
        existing_usd = existing_usd[existing_usd["COUNTRY"].astype(str).str.upper() != country_upper]
    write_global_csv(existing_usd, GLOBAL_DATA_CSV)

    existing_local = read_global_csv(GLOBAL_LOCAL_DATA_CSV)
    if not existing_local.empty and "COUNTRY" in existing_local.columns:
        existing_local = existing_local[existing_local["COUNTRY"].astype(str).str.upper() != country_upper]
    write_global_csv(existing_local, GLOBAL_LOCAL_DATA_CSV)

    meta = load_upload_metadata()
    if country_upper in meta:
        del meta[country_upper]
    save_upload_metadata(meta)

    if not existing_usd.empty:
        save_normalized_excel(existing_usd, GLOBAL_XLSX, collect_metadata_rows_from_meta(meta), 0.0)
    elif GLOBAL_XLSX.exists():
        GLOBAL_XLSX.unlink()

    for path in (ORIGINALS_DIR / f"{country}_original.xlsx", LOCAL_DIR / f"{country}.xlsx", USD_DIR / f"{country}.xlsx"):
        if path.exists():
            path.unlink()

    refresh_dashboard_data()

# ==========================================================
# UPLOAD PIPELINE
# ==========================================================
def process_upload(
    file_bytes: bytes,
    source_file_name: str,
    sheet_name: str,
    country: str,
    input_currency: str,
    money_factor_label: str,
    pairs_factor_label: str,
    money_columns: list[str],
    pairs_columns: list[str],
    vat_input_percent: float,
) -> dict:
    money_factor = FACTOR_OPTIONS[money_factor_label]
    pairs_factor = FACTOR_OPTIONS[pairs_factor_label]
    vat_rate = float(vat_input_percent) / 100.0
    fx_rate = get_fx_rate(country)

    detail_rows = read_uploaded_workbook(file_bytes, sheet_name)
    df_lc, df_usd = normalize_source_rows(
        detail_rows=detail_rows,
        money_factor=money_factor,
        pairs_factor=pairs_factor,
        money_columns=money_columns,
        pairs_columns=pairs_columns,
        input_currency=input_currency,
        fx_rate=fx_rate,
        country=country,
    )
    df_lc = aggregate_duplicates(df_lc)
    df_usd = aggregate_duplicates(df_usd)
    if df_usd.empty:
        raise ValueError("No valid detail rows were found in the uploaded file.")

    metadata_row = {
        "COUNTRY": country,
        "UPLOAD_DATE": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "SOURCE_FILE": source_file_name,
        "CURRENCY": input_currency,
        "FX_RATE": fx_rate,
        "MONEY_FACTOR": money_factor_label,
        "PAIRS_FACTOR": pairs_factor_label,
        "VAT_RATE": vat_rate,
        "ROW_COUNT": int(len(df_usd)),
        "vat_rate": vat_rate,
        "WORKSHEET": sheet_name,
    }

    original_path = ORIGINALS_DIR / f"{country}_original.xlsx"
    local_path = LOCAL_DIR / f"{country}.xlsx"
    usd_path = USD_DIR / f"{country}.xlsx"
    targets = [
        original_path, local_path, usd_path,
        GLOBAL_DATA_CSV, GLOBAL_LOCAL_DATA_CSV, GLOBAL_XLSX, UPLOAD_META_PATH,
    ]
    backups = {path: _backup_file(path) for path in targets}
    created: list[Path] = []
    try:
        existing_usd = read_global_csv(GLOBAL_DATA_CSV)
        existing_local = read_global_csv(GLOBAL_LOCAL_DATA_CSV)
        combined_usd = upsert_country_frames(existing_usd, df_usd, country)
        combined_local = upsert_country_frames(existing_local, df_lc, country)

        original_path.write_bytes(file_bytes)
        created.append(original_path)
        save_normalized_excel(df_lc, local_path, [metadata_row], vat_rate)
        created.append(local_path)
        save_normalized_excel(df_usd, usd_path, [metadata_row], vat_rate)
        created.append(usd_path)

        meta = load_upload_metadata()
        meta[country.upper()] = metadata_row
        save_upload_metadata(meta)
        created.append(UPLOAD_META_PATH)

        write_global_csv(combined_usd, GLOBAL_DATA_CSV)
        created.append(GLOBAL_DATA_CSV)
        write_global_csv(combined_local, GLOBAL_LOCAL_DATA_CSV)
        created.append(GLOBAL_LOCAL_DATA_CSV)
        save_normalized_excel(
            combined_usd,
            GLOBAL_XLSX,
            collect_metadata_rows(country, metadata_row),
            vat_rate,
        )
        created.append(GLOBAL_XLSX)
        _cleanup_backups(backups)
    except Exception:
        _restore_outputs(backups, created)
        _cleanup_backups(backups)
        raise

    return {
        "country": country,
        "row_count": int(len(df_usd)),
        "currency": input_currency,
        "fx_rate": fx_rate,
        "vat_rate": vat_rate,
    }

def build_preview_dataframe(
    detail_rows: list[dict],
    money_factor: float,
    pairs_factor: float,
    money_columns: list[str],
    pairs_columns: list[str],
    input_currency: str = "USD",
    fx_rate: float = 1.0,
) -> pd.DataFrame:
    id_cols = ["KEY", "BU", "GENDER", "OCCASION", "CATEGORY"]
    value_cols = [h for h in HEADERS if "VALUE" in h and h not in OMIT_COLUMNS]
    pairs_set = set(pairs_columns)
    money_set = set(money_columns)
    apply_fx = input_currency != "USD" and fx_rate not in (None, 0)
    records = []
    for row in detail_rows[:15]:
        rec = {col: row.get(col, "") for col in id_cols}
        for header in value_cols:
            raw = parse_number(row.get(header))
            rec[header] = raw
            scaled = raw
            if scaled is not None:
                if header in money_set:
                    scaled = scaled * money_factor
                if header in pairs_set:
                    scaled = scaled * pairs_factor
                if apply_fx and header in MONEY_COLUMNS:
                    scaled = scaled / fx_rate
            rec[f"{header} (scaled)"] = scaled
        records.append(rec)
    ordered = id_cols[:]
    for header in value_cols:
        ordered.extend([header, f"{header} (scaled)"])
    return pd.DataFrame(records, columns=ordered)

# ==========================================================
# DASHBOARD - LOAD GLOBAL DATA
# ==========================================================
@st.cache_data(show_spinner=False)
def load_global_csv_data(csv_path_text: str) -> pd.DataFrame:
    path = Path(csv_path_text)
    if not path.exists():
        return pd.DataFrame()
    data = pd.read_csv(
        path,
        sep=";",
        decimal=",",
        encoding="utf-8-sig",
    )
    data.columns = [str(c).replace("\ufeff", "").strip() for c in data.columns]
    for column in list(data.columns):
        if is_text_column(column) or is_separator_column(column) or is_omitted_column(column):
            data[column] = data[column].map(clean_text)
        else:
            data[column] = pd.to_numeric(data[column], errors="coerce")
    return data

def dashboard_source_available() -> bool:
    return GLOBAL_DATA_CSV.exists() or GLOBAL_LOCAL_DATA_CSV.exists()

def get_dashboard_dataset(mode: str) -> pd.DataFrame:
    if mode == "Local Currency":
        path = GLOBAL_LOCAL_DATA_CSV
    else:
        path = GLOBAL_DATA_CSV
    return load_global_csv_data(str(path)).copy()

def validate_dashboard_columns(data: pd.DataFrame) -> list[str]:
    required = [
        "BU", "GENDER", "OCCASION", "CATEGORY",
        "TURNOVER VALUE 2025", "TURNOVER VALUE 2026", "TURNOVER VALUE 2027",
        "PAIRS VALUE 2025", "PAIRS VALUE 2026", "PAIRS VALUE 2027",
        "MKD VALUE 2026", "MKD VALUE 2027",
        "GROSS VALUE 2026", "GROSS VALUE 2027",
        "MRG VALUE 2026", "MRG VALUE 2027",
        "COUNTRY",
    ]
    return [c for c in required if find_column(data, c) is None]

def refresh_dashboard_data() -> None:
    load_global_csv_data.clear()

# ==========================================================
# DASHBOARD - FILTERING
# ==========================================================
FILTER_ORDER = ["COUNTRY", "BU", "GENDER", "OCCASION", "CATEGORY"]
HIERARCHY_LEVELS = ["BU", "GENDER", "OCCASION", "CATEGORY"]

def filter_data_by_values(data: pd.DataFrame, selections: dict) -> pd.DataFrame:
    filtered = data
    for column, selected_values in selections.items():
        if selected_values and column in filtered.columns:
            filtered = filtered[filtered[column].isin(selected_values)]
    return filtered

def get_filter_options(data: pd.DataFrame, column: str, current_selections: dict) -> list[str]:
    if column not in data.columns:
        return []
    partial = data
    for previous_column in FILTER_ORDER[: FILTER_ORDER.index(column)]:
        selected_values = current_selections.get(previous_column, [])
        if selected_values and previous_column in partial.columns:
            partial = partial[partial[previous_column].isin(selected_values)]
    values = [clean_text(v) for v in partial[column].tolist() if clean_text(v)]
    return sorted(dict.fromkeys(values))

def make_filter_multiselect(label: str, column: str, data: pd.DataFrame, selections: dict, session_key: str):
    options = get_filter_options(data, column, selections)
    return st.multiselect(label, options=options, key=session_key)

# ==========================================================
# DASHBOARD - CALCULATIONS
# ==========================================================
def numeric_series(data: pd.DataFrame, column_name: str) -> pd.Series:
    if column_name not in data.columns:
        return pd.Series([0.0] * len(data), index=data.index, dtype=float)
    return pd.to_numeric(data[column_name], errors="coerce")

def numeric_sum(data: pd.DataFrame, column_name: str) -> float:
    if data.empty or column_name not in data.columns:
        return 0.0
    return float(pd.to_numeric(data[column_name], errors="coerce").fillna(0.0).sum())

def recalculate_calculated_columns(data: pd.DataFrame, vat_rate: float) -> pd.DataFrame:
    out = data.copy()
    f = numeric_series(out, "TURNOVER VALUE 2025")
    g = numeric_series(out, "TURNOVER VALUE 2026")
    h = numeric_series(out, "TURNOVER VALUE 2027")
    aw = numeric_series(out, "MKD VALUE 2025")
    ax = numeric_series(out, "MKD VALUE 2026")
    ay = numeric_series(out, "MKD VALUE 2027")
    ba = numeric_series(out, "MRG VALUE 2025")
    bb = numeric_series(out, "MRG VALUE 2026")
    bc = numeric_series(out, "MRG VALUE 2027")
    be = numeric_series(out, "COGS VALUE 2025")
    bf = numeric_series(out, "COGS VALUE 2026")
    bg = numeric_series(out, "COGS VALUE 2027")
    bi = numeric_series(out, "GROSS VALUE 2025")
    bj = numeric_series(out, "GROSS VALUE 2026")
    bk = numeric_series(out, "GROSS VALUE 2027")
    bm = numeric_series(out, "PAIRS VALUE 2025")
    bn = numeric_series(out, "PAIRS VALUE 2026")
    bo = numeric_series(out, "PAIRS VALUE 2027")
    bv = numeric_series(out, "INITIAL STOCK QTY VALUE 2025 H1")
    bw = numeric_series(out, "FINAL STOCK QTY VALUE 2025 H1")
    by = numeric_series(out, "INITIAL STOCK QTY VALUE 2026 H1")
    bz = numeric_series(out, "FINAL STOCK QTY VALUE 2026 H1")
    cb = numeric_series(out, "INITIAL STOCK QTY VALUE 2027 H1")
    cc = numeric_series(out, "FINAL STOCK QTY VALUE 2027 H1")
    cf = numeric_series(out, "OTB QTY 2026 H1")
    ci = numeric_series(out, "OTB QTY 2027 H1")

    out["TO TALY% VS 2025"] = series_div(h, f)
    out["TO TALY% VS 2026"] = series_div(h, g)
    out["TO CONTRIBUTION % 2025"] = series_div(f, f.sum())
    out["TO CONTRIBUTION % 2026"] = series_div(g, g.sum())
    out["TO CONTRIBUTION % 2027"] = series_div(h, h.sum())
    out["INTAKE% 2025"] = series_div(bi - be, bi)
    out["INTAKE% 2026"] = series_div(bj - bf, bj)
    out["INTAKE% 2027"] = series_div(bk - bg, bk)
    out["INTAKE TALY% VS 2025"] = out["INTAKE% 2027"] - out["INTAKE% 2025"]
    out["INTAKE TALY% VS 2026"] = out["INTAKE% 2027"] - out["INTAKE% 2026"]
    out["MKD% 2025"] = series_div(aw, bi)
    out["MKD% 2026"] = series_div(ax, bj)
    out["MKD% 2027"] = series_div(ay, bk)
    out["MRG% 2025"] = series_div(ba, f)
    out["MRG% 2026"] = series_div(bb, g)
    out["MRG% 2027"] = series_div(bc, h)
    out["RRP 2025"] = series_div(bi, bm)
    out["RRP 2026"] = series_div(bj, bn)
    out["RRP 2027"] = series_div(bk, bo)
    out["RRP TALY% VS 2025"] = series_div(out["RRP 2027"], out["RRP 2025"])
    out["RRP TALY% VS 2026"] = series_div(out["RRP 2027"], out["RRP 2026"])
    out["RRP VAT 2025"] = out["RRP 2025"] * (1 + vat_rate)
    out["RRP VAT 2026"] = out["RRP 2026"] * (1 + vat_rate)
    out["RRP VAT 2027"] = out["RRP 2027"] * (1 + vat_rate)
    out["RRP VAT TALY% VS 2025"] = series_div(out["RRP VAT 2027"], out["RRP VAT 2025"])
    out["RRP VAT TALY% VS 2026"] = series_div(out["RRP VAT 2027"], out["RRP VAT 2026"])
    out["ASP 2025"] = series_div(f, bm)
    out["ASP 2026"] = series_div(g, bn)
    out["ASP 2027"] = series_div(h, bo)
    out["ASP TALY% VS 2025"] = series_div(out["ASP 2027"], out["ASP 2025"])
    out["ASP TALY% VS 2026"] = series_div(out["ASP 2027"], out["ASP 2026"])
    out["PAIRS TALY% VS 2025"] = series_div(bo, bm)
    out["PAIRS TALY% VS 2026"] = series_div(bo, bn)
    out["PAIRS CONTRIBUTION % 2025"] = series_div(bm, bm.sum())
    out["PAIRS CONTRIBUTION % 2026"] = series_div(bn, bn.sum())
    out["PAIRS CONTRIBUTION % 2027"] = series_div(bo, bo.sum())
    out["% FINAL STOCK VS INITIAL 25H1"] = series_div(bw, bv)
    out["% FINAL STOCK VS INITIAL 26H1"] = series_div(bz, by)
    out["% FINAL STOCK VS INITIAL 27H1"] = series_div(cc, cb)
    out["SOR FINAL 2026 H2"] = series_div(bn, bn + bz)
    out["SOR TARGET 2027H1"] = series_div(cc, ci)
    out["STOCK TURN 2025"] = series_div(bm, (bv + bw) / 2.0)
    out["STOCK TURN 2026"] = series_div(bn, (by + bz) / 2.0)
    out["STOCK TURN 2027"] = series_div(bo, (cb + cc) / 2.0)

    if "KEY" not in out.columns:
        out["KEY"] = ""
    bu = out["BU"].map(clean_text) if "BU" in out.columns else ""
    gender = out["GENDER"].map(clean_text) if "GENDER" in out.columns else ""
    occasion = out["OCCASION"].map(clean_text) if "OCCASION" in out.columns else ""
    category = out["CATEGORY"].map(clean_text) if "CATEGORY" in out.columns else ""
    if isinstance(bu, pd.Series):
        reconstructed = bu + gender + occasion + category
        missing = out["KEY"].map(clean_text) == ""
        out.loc[missing, "KEY"] = reconstructed[missing]
    return out

def calculate_dashboard_metrics(data: pd.DataFrame) -> dict:
    t26 = numeric_sum(data, "TURNOVER VALUE 2026")
    t27 = numeric_sum(data, "TURNOVER VALUE 2027")
    p26 = numeric_sum(data, "PAIRS VALUE 2026")
    p27 = numeric_sum(data, "PAIRS VALUE 2027")
    mrg_26 = numeric_sum(data, "MRG VALUE 2026")
    mrg_27 = numeric_sum(data, "MRG VALUE 2027")
    mkd_26 = numeric_sum(data, "MKD VALUE 2026")
    mkd_27 = numeric_sum(data, "MKD VALUE 2027")
    gross_26 = numeric_sum(data, "GROSS VALUE 2026")
    gross_27 = numeric_sum(data, "GROSS VALUE 2027")

    asp_2027 = safe_divide(t27, p27)
    asp_2026 = safe_divide(t26, p26)
    mkd_pct_2027 = safe_divide(mkd_27, gross_27)
    mkd_pct_2026 = safe_divide(mkd_26, gross_26)
    mrg_pct_2027 = safe_divide(mrg_27, t27)
    mrg_pct_2026 = safe_divide(mrg_26, t26)

    return {
        "turnover_2027": t27,
        "turnover_taly_2026": calc_growth(t27, t26),
        "pairs_2027": p27,
        "pairs_taly_2026": calc_growth(p27, p26),
        "asp_2027": asp_2027,
        "asp_taly_2026": calc_growth(asp_2027, asp_2026),
        "margin_pct_2027": mrg_pct_2027,
        "margin_delta_2026": mrg_pct_2027 - mrg_pct_2026,
        "markdown_pct_2027": mkd_pct_2027,
        "markdown_delta_2026": mkd_pct_2027 - mkd_pct_2026,
    }

def render_metric_block(title: str, metrics: dict, currency_label: str) -> None:
    st.markdown(f'<div class="dashboard-section-title">{title}</div>', unsafe_allow_html=True)
    cols = st.columns(5)
    cols[0].metric(
        f"Turnover 2027 ({currency_label})",
        format_integer(metrics["turnover_2027"]),
        delta=format_delta_kpi(metrics["turnover_taly_2026"]),
    )
    cols[1].metric(
        "Volume 2027",
        format_integer(metrics["pairs_2027"]),
        delta=format_delta_kpi(metrics["pairs_taly_2026"]),
    )
    cols[2].metric(
        f"ASP 2027 ({currency_label})",
        format_decimal(metrics["asp_2027"], 2),
        delta=format_delta_kpi(metrics["asp_taly_2026"]),
    )
    cols[3].metric(
        "Margin 2027",
        format_percent(metrics["margin_pct_2027"], 0),
        delta=format_pp_delta(metrics["margin_delta_2026"]),
    )
    cols[4].metric(
        "Markdown 2027",
        format_percent(metrics["markdown_pct_2027"], 0),
        delta=format_pp_delta(metrics["markdown_delta_2026"]),
    )

# ==========================================================
# DASHBOARD - TABLES & CHARTS
# ==========================================================
def build_hierarchy_summary(data: pd.DataFrame, hierarchy_level: str, vat_rate: float) -> pd.DataFrame:
    if data.empty or hierarchy_level not in HIERARCHY_LEVELS:
        hierarchy_level = "BU"
    levels_to_group = [c for c in HIERARCHY_LEVELS[: HIERARCHY_LEVELS.index(hierarchy_level) + 1] if c in data.columns]
    if not levels_to_group:
        return pd.DataFrame()
    working = data.copy()
    additive = [c for c in BASE_ADDITIVE_COLUMNS if c in working.columns]
    working[additive] = working[additive].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    grouped = working.groupby(levels_to_group, as_index=False, dropna=False, sort=False)[additive].sum()
    grouped = recalculate_calculated_columns(grouped, vat_rate)
    remaining = [
        c for c in HEADERS
        if c not in TEXT_COLUMNS
        and c not in OMIT_COLUMNS
        and not is_separator_column(c)
    ]
    return grouped.reindex(columns=levels_to_group + remaining)

def build_share_data(data: pd.DataFrame, hierarchy_level: str, value_column: str) -> pd.DataFrame:
    if data.empty or hierarchy_level not in data.columns or value_column not in data.columns:
        return pd.DataFrame()
    chart_data = data[[hierarchy_level, value_column]].copy()
    chart_data[value_column] = pd.to_numeric(chart_data[value_column], errors="coerce").fillna(0.0)
    chart_data = (
        chart_data[chart_data[value_column] > 0]
        .groupby(hierarchy_level, as_index=False, sort=False)[value_column]
        .sum()
        .sort_values(value_column, ascending=False)
    )
    if len(chart_data) > 12:
        chart_data = pd.concat(
            [
                chart_data.head(11),
                pd.DataFrame({hierarchy_level: ["Other"], value_column: [chart_data.iloc[11:][value_column].sum()]}),
            ],
            ignore_index=True,
        )
    return chart_data

def render_share_chart(data: pd.DataFrame, hierarchy_level: str, value_column: str, title: str) -> None:
    chart_data = build_share_data(data, hierarchy_level, value_column)
    if chart_data.empty:
        st.info(f"No data available to build {title}.")
        return
    figure = px.pie(
        chart_data,
        names=hierarchy_level,
        values=value_column,
        hole=0.38,
        color_discrete_sequence=px.colors.qualitative.Set3,
    )
    figure.update_traces(
        textposition="inside",
        textinfo="percent+label",
        hovertemplate="<b>%{label}</b>\nValue: %{value:,.0f}\nShare: %{percent:.1%}<extra></extra>",
    )
    figure.update_layout(
        title=dict(text=title, x=0.5, xanchor="center"),
        height=460,
        margin=dict(l=10, r=10, t=55, b=10),
        legend=dict(title_text=hierarchy_level),
    )
    st.plotly_chart(figure, width="stretch", config={"displaylogo": False, "responsive": True})

def render_combined_chart(data: pd.DataFrame, title: str, bar_label: str,
                          bar_values: list[float], line_label: str,
                          line_values: list[float], line_is_percent: bool) -> None:
    years = ["2025", "2026", "2027"]

    fig = go.Figure()

    fig.add_trace(
        go.Bar(
            x=years,
            y=bar_values,
            name=bar_label,
            marker_color="#1F6F78",
            text=[format_integer(v) for v in bar_values],
            textposition="outside",
            yaxis="y1",
        )
    )

    line_text = (
        [format_percent(v, 0) for v in line_values]
        if line_is_percent
        else [format_decimal(v, 2) for v in line_values]
    )

    fig.add_trace(
        go.Scatter(
            x=years,
            y=line_values,
            name=line_label,
            mode="lines+markers+text",
            text=line_text,
            textposition="top center",
            line=dict(color="#E8792F", width=3),
            marker=dict(size=8),
            yaxis="y2",
        )
    )

    fig.update_layout(
        title=dict(text=title, x=0.5, xanchor="center"),
        height=420,
        margin=dict(l=10, r=10, t=55, b=10),
        xaxis=dict(
            type="category",
            categoryorder="array",
            categoryarray=["2025", "2026", "2027"],
            tickmode="array",
            tickvals=["2025", "2026", "2027"],
            ticktext=["2025", "2026", "2027"],
        ),
        yaxis=dict(title=bar_label, side="left"),
        yaxis2=dict(
            title=line_label,
            overlaying="y",
            side="right",
            tickformat=".0%" if line_is_percent else None,
        ),
        legend=dict(orientation="h", yanchor="bottom", y=-0.2),
    )

    st.plotly_chart(fig, width="stretch", config={"displaylogo": False, "displayModeBar": False})

def insert_separator_columns(columns: list[str], order: list) -> list[str]:
    selected = [c for c in columns if c and not is_separator_column(c)]
    selected_set = set(selected)
    groups: list[list[str]] = []
    current: list[str] = []
    for item in order:
        if item is None:
            if current:
                groups.append(current)
            current = []
            continue
        if item in selected_set:
            current.append(item)
    if current:
        groups.append(current)
    placed = {c for g in groups for c in g}
    extras = [c for c in selected if c not in placed]
    if extras:
        groups.append(extras)
    result: list[str] = []
    for index, group in enumerate(groups):
        if index:
            result.append(make_sep_name(index))
        result.extend(group)
    return result

def _taly_color(value, midpoint: float) -> str:
    try:
        if value is None or pd.isna(value):
            return ""
        number = float(value)
    except (TypeError, ValueError):
        return ""
    if number > midpoint:
        return "background-color: #D8EEDC; color: #1b1b1b"
    if abs(number - midpoint) < 1e-9:
        return "background-color: #EEF7EF; color: #1b1b1b"
    return "background-color: #F6DDE0; color: #1b1b1b"

def style_dashboard_table(data: pd.DataFrame):
    formatted = data.copy()
    separators = [c for c in formatted.columns if is_separator_column(c)]
    text_columns = [c for c in formatted.columns if is_text_column(c)]
    ratio_cf = [c for c in formatted.columns if c in RATIO_TALY_CF_COLUMNS]
    intake_cf = [c for c in formatted.columns if c in INTAKE_TALY_CF_COLUMNS]

    def formatter_for(column: str):
        if is_separator_column(column) or is_omitted_column(column):
            return lambda v: ""
        if is_text_column(column):
            return lambda v: clean_text(v)
        if is_percentage_column(column):
            return lambda v: format_percent(v, 0)
        if is_decimal_metric_column(column):
            return lambda v: format_decimal(v, 2)
        if is_count_or_money_column(column):
            return lambda v: format_integer(v)
        return lambda v: format_decimal(v, 2)

    styler = formatted.style.set_table_styles(
        [
            {"selector": "thead th", "props": [
                ("background-color", "#252525"), ("color", "#ffffff"), ("font-weight", "700"),
                ("border", "1px solid #ffffff"), ("text-align", "center"), ("font-size", "11px"),
                ("white-space", "normal"), ("min-width", "0px"),
            ]},
            {"selector": "tbody td", "props": [
                ("border", "1px solid #e6e9ef"), ("font-size", "11px"), ("padding", "5px 7px"),
            ]},
            {"selector": "tbody tr:nth-child(even)", "props": [
                ("background-color", "#fafbfc"),
            ]},
        ]
    )
    
    if ratio_cf:
        styler = styler.apply(lambda col: col.map(lambda v: _taly_color(v, 1.0)), subset=ratio_cf)
    if intake_cf:
        styler = styler.apply(lambda col: col.map(lambda v: _taly_color(v, 0.0)), subset=intake_cf)
        
    styler = styler.format({column: formatter_for(column) for column in formatted.columns})

    if text_columns:
        styler = styler.set_properties(subset=text_columns, **{"text-align": "left", "white-space": "nowrap"})
        
    dimension_columns = [
        c for c in formatted.columns
        if normalize_column_name(c) in {"BU", "GENDER", "OCCASION", "CATEGORY"}
    ]
    if dimension_columns:
        styler = styler.set_properties(
            subset=dimension_columns,
            **{
                "font-weight": "700",
                "color": "#1B1B1B",
            },
        )
        
    numeric_columns = [
        c for c in formatted.columns 
        if c not in text_columns 
        and c not in separators 
    ]
    if numeric_columns:
        styler = styler.set_properties(subset=numeric_columns, **{"text-align": "right", "white-space": "nowrap"})

    if separators:
        styler = styler.set_properties(
            subset=separators,
            **{
                "background-color": "#6B7280", "color": "#6B7280", "width": "5px",
                "min-width": "5px", "max-width": "5px", "padding": "0px",
                "border": "none",
            },
        )
    return styler

def build_column_config(columns) -> dict:
    pinned = {"BU", "GENDER", "OCCASION", "CATEGORY"}
    config = {}
    for column in columns:
        if is_separator_column(column):
            # Keep the thin separator columns as-is; only label/width of real columns change.
            config[column] = st.column_config.TextColumn("", width=5)
            continue
        label = display_header_label(column)
        width = max(65, min(125, len(label) * 8 + 18))
        config[column] = st.column_config.TextColumn(
            label,
            width=width,
            pinned=normalize_column_name(column) in pinned,
        )
    return config

def render_refresh_button() -> None:
    if st.button(
        "REFRESH DASHBOARD DATA",
        key="refresh_dashboard_data",
        width="stretch",
    ):
        refresh_dashboard_data()
        st.rerun()

def render_download_button(
    label: str,
    data_bytes: bytes | None,
    file_name: str,
    empty_caption: str,
) -> None:
    if data_bytes:
        st.download_button(
            label,
            data=data_bytes,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
    else:
        st.caption(empty_caption)

# ==========================================================
# PAGES
# ==========================================================
def show_page_header(active_page: str) -> None:
    header_left, header_spacer, header_upload, header_visualization, header_sep, header_logout = st.columns(
        [1.35, 5.65, 1.20, 1.20, 0.08, 1.15],
        vertical_alignment="center",
    )

    with header_left:
        if LOGO_PATH.exists():
            st.image(str(LOGO_PATH), width=110)

    with header_upload:
        st.markdown(
            f'<div class="nav-button-slot {"active" if active_page == "upload" else ""}">',
            unsafe_allow_html=True,
        )
        if st.button("UPLOAD FILE", key="nav_upload", width="stretch"):
            st.session_state["page"] = "upload"
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    with header_visualization:
        st.markdown(
            f'<div class="nav-button-slot {"active" if active_page == "dashboard" else ""}">',
            unsafe_allow_html=True,
        )
        dashboard_disabled = not dashboard_source_available()
        if st.button(
            "VISUALIZATION",
            key="nav_dashboard",
            width="stretch",
            disabled=dashboard_disabled,
        ):
            st.session_state["page"] = "dashboard"
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    with header_sep:
        st.markdown('<div class="logout-separator"></div>', unsafe_allow_html=True)

    with header_logout:
        st.markdown('<div class="nav-button-slot">', unsafe_allow_html=True)
        if st.button("LOG OUT", key=f"logout_top_{active_page}", width="stretch"):
            st.session_state.clear()
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="top-header-divider"></div>', unsafe_allow_html=True)

def show_login() -> None:
    login_width_px = 520
    st.markdown(
        f"""
        <style>
        div[data-testid="stForm"] {{
            max-width: {login_width_px}px;
            width: 100%;
            margin: 6vh auto 0 auto;
            padding: 0 !important;
            border: none !important;
        }}
        div[data-testid="stForm"] [data-testid="stImage"],
        div[data-testid="stForm"] [data-testid="stImage"] > div {{
            display: flex !important;
            justify-content: center !important;
            width: 100% !important;
            position: relative !important;
            left: 73px !important;
        }}
        div[data-testid="stForm"] img {{
            display: block !important;
            margin: 0 auto !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )
    with st.form("login_form", border=False):
        if LOGO_PATH.exists():
            st.image(str(LOGO_PATH), width=220)
        st.markdown(
            "<h3 style='text-align:center;margin:0 0 1.2rem 0;font-weight:700'>Global Category Plan</h3>",
            unsafe_allow_html=True,
        )
        username = st.selectbox(
            "Username", options=MENU_OPTIONS, index=None,
            placeholder="Select your username", key="login_username_select",
        )
        password = st.text_input("Password", type="password", key="login_password_input")
        submitted = st.form_submit_button("SIGN IN", width="stretch")
        if submitted:
            if not username:
                st.error("Please select a username.")
                return
            user = USERS.get(username)
            if user is None or password != user["password"]:
                st.error("Invalid password.")
                return
            st.session_state.update(
                authenticated=True, username=username, country=user["country"],
                role=user["role"], page="upload",
            )
            st.rerun()

def show_upload_page() -> None:
    show_page_header("upload")
    role = st.session_state["role"]
    assigned_country = st.session_state["country"]

    st.title("DATA UPLOAD & TRANSFORMATION")
    if st.session_state.get("last_upload_message"):
        st.success(st.session_state["last_upload_message"])

    file_bytes = st.session_state.get("upload_file_bytes")
    source_file_name = st.session_state.get("upload_file_name")

    # ---- File, worksheet and country ------------------------------------
    with st.container(border=True):
        uploaded_file = st.file_uploader(
            "UPLOAD EXCEL FILE",
            type=["xlsx", "xlsm"],
            help="Use the category plan template in .xlsx or .xlsm format.",
            key="uploaded_excel_file",
        )
        if uploaded_file is not None:
            st.session_state["upload_file_bytes"] = uploaded_file.getvalue()
            st.session_state["upload_file_name"] = uploaded_file.name
            file_bytes = st.session_state["upload_file_bytes"]
            source_file_name = st.session_state["upload_file_name"]

        sheet_names: list[str] = []
        if file_bytes:
            try:
                sheet_names = workbook_sheet_names(file_bytes)
            except Exception as error:
                st.error(f"The Excel file could not be read: {error}")
                return

        default_index = (
            sheet_names.index(SOURCE_SHEET_NAME)
            if sheet_names and SOURCE_SHEET_NAME in sheet_names
            else 0
        )

        col_sheet, col_country = st.columns(2)
        with col_sheet:
            if sheet_names:
                selected_sheet = st.selectbox(
                    "SELECT WORKSHEET",
                    sheet_names,
                    index=default_index,
                    key="selected_upload_sheet",
                )
            else:
                selected_sheet = None
        with col_country:
            if role == "admin":
                target_country = st.selectbox(
                    "TARGET COUNTRY",
                    options=COUNTRY_USERS,
                    key="upload_target_country",
                )
            else:
                target_country = assigned_country
                st.info(f"Target country: {target_country.title()}")

    if not file_bytes:
        st.caption("Choose an Excel file to review and transform.")
        return

    if not selected_sheet:
        st.error("Please select a worksheet.")
        return

    detail_rows = []
    try:
        validate_selected_worksheet(file_bytes, selected_sheet)
        detail_rows = read_uploaded_workbook(file_bytes, selected_sheet)
    except Exception as exc:
        st.error(str(exc))
        detail_rows = []

    if not detail_rows:
        st.error("No valid detail rows were found. Check the selected worksheet, B7:E7 labels, and that data starts at row 8.")
        return

    # ---- Transformation options -----------------------------------------
    st.subheader("⚙️ TRANSFORMATION OPTIONS")

    if role == "admin":
        currency_options = ALL_CURRENCY_OPTIONS
    else:
        currency_options = COUNTRY_CURRENCIES.get(target_country, ["USD"])

    st.session_state.setdefault("currency", currency_options[0])
    st.session_state.setdefault("money_factor", "/1")
    st.session_state.setdefault("pairs_factor", "/1")
    suggested_vat = DEFAULT_VAT_RATES.get(target_country.upper(), 0.0)
    if st.session_state.get("vat_input_country") != target_country:
        st.session_state["vat_input"] = suggested_vat
        st.session_state["vat_input_country"] = target_country
    st.session_state.setdefault("vat_input", suggested_vat)
    st.session_state.setdefault("money_columns", list(MONEY_COLUMNS))
    st.session_state.setdefault("pairs_columns", list(PAIRS_SCALABLE_COLUMNS))
    if st.session_state.get("currency") not in currency_options:
        st.session_state["currency"] = currency_options[0]

    # General settings (Currency and VAT)
    with st.container(border=True):
        col_curr, col_vat = st.columns(2)
        with col_curr:
            input_currency = st.selectbox(
                "CURRENCY OF UPLOADED FILE",
                currency_options,
                key="currency",
            )
        with col_vat:
            vat_input_percent = st.number_input(
                "VAT (IVA) RATE %",
                min_value=0.0,
                max_value=100.0,
                step=0.5,
                key="vat_input",
                help="Used to calculate RRP VAT only. Suggested rate based on target country; adjust if needed.",
            )

    # Money and Pairs scaling in two parallel columns
    col_money, col_pairs = st.columns(2)

    with col_money:
        with st.container(border=True):
            st.markdown("#### 💵 MONEY VALUE SCALING")
            money_factor_label = st.selectbox(
                "MONEY VALUE SCALING FACTOR",
                FACTOR_OPTIONS_LIST,
                format_func=lambda x: "No division" if x in ["/1", "*1"] else x,
                key="money_factor",
            )
            money_columns = st.multiselect(
                "COLUMNS AFFECTED BY MONEY SCALING FACTOR",
                options=MONEY_COLUMNS,
                default=list(MONEY_COLUMNS),
                key="money_columns",
            )

    with col_pairs:
        with st.container(border=True):
            st.markdown("#### 📦 PAIRS / UNITS SCALING")
            pairs_factor_label = st.selectbox(
                "PAIRS/UNITS SCALING FACTOR",
                FACTOR_OPTIONS_LIST,
                format_func=lambda x: "No division" if x in ["/1", "*1"] else x,
                key="pairs_factor",
            )
            pairs_columns = st.multiselect(
                "COLUMNS AFFECTED BY PAIRS/UNITS SCALING FACTOR",
                options=PAIRS_SCALABLE_COLUMNS,
                default=list(PAIRS_SCALABLE_COLUMNS),
                key="pairs_columns",
            )

    # ---- Preview ---------------------------------------------------------
    try:
        preview_fx_rate = get_fx_rate(target_country) if input_currency != "USD" else 1.0
    except ValueError:
        preview_fx_rate = 1.0

    preview_df = build_preview_dataframe(
        detail_rows=detail_rows,
        money_factor=FACTOR_OPTIONS[money_factor_label],
        pairs_factor=FACTOR_OPTIONS[pairs_factor_label],
        money_columns=money_columns,
        pairs_columns=pairs_columns,
        input_currency=input_currency,
        fx_rate=preview_fx_rate,
    )

    st.subheader("PREVIEW")
    if input_currency != "USD":
        st.caption(f"PREVIEW {input_currency} → USD (Tasa {preview_fx_rate:.3f})")
    else:
        st.caption("PREVIEW USD — currency conversion is not applied.")

    with st.expander("Upload diagnostics", expanded=False):
        st.write("Detected detail rows:", len(detail_rows))
        st.write("Preview rows:", len(preview_df))
        st.write("Selected worksheet:", selected_sheet)

    if preview_df.empty:
        st.warning("No preview rows were generated.")
    else:
        scaled_cols = [c for c in preview_df.columns if c.endswith("(scaled)")]
        numeric_cols = [
            c for c in preview_df.columns
            if c not in ("KEY", "BU", "GENDER", "OCCASION", "CATEGORY")
        ]
        preview_styler = (
            preview_df.style
            .format({col: format_integer for col in numeric_cols})
            .set_properties(subset=scaled_cols, **{"background-color": "#EDEDED"})
            .set_properties(**{"text-align": "center"})
        )
        st.table(preview_styler)

    load_clicked = st.button(
        "LOAD DATA",
        type="primary",
        use_container_width=True,
        key="load_data_button",
    )
    if load_clicked:
        try:
            result = process_upload(
                file_bytes=file_bytes,
                source_file_name=source_file_name or (uploaded_file.name if uploaded_file is not None else "upload.xlsx"),
                sheet_name=selected_sheet,
                country=target_country,
                input_currency=input_currency,
                money_factor_label=money_factor_label,
                pairs_factor_label=pairs_factor_label,
                money_columns=money_columns,
                pairs_columns=pairs_columns,
                vat_input_percent=vat_input_percent,
            )
            refresh_dashboard_data()
            st.session_state["last_upload_message"] = (
                f"File processed successfully for {result['country']}. "
                f"Rows loaded: {result['row_count']}."
            )
            st.session_state["page"] = "dashboard"
            st.rerun()
        except Exception:
            st.error("Your file could not be processed. Please try again.")
            with st.expander("Technical details", expanded=False):
                st.code(traceback.format_exc())

def show_dashboard_page() -> None:
    show_page_header("dashboard")
    role = st.session_state["role"]
    assigned_country = st.session_state["country"]

    st.markdown('<div class="dashboard-main-title">Global Category Plan 2027 H2</div>', unsafe_allow_html=True)

    if not dashboard_source_available():
        st.warning("No consolidated global file was found.")
        return

    st.session_state.setdefault("dashboard_currency_mode", "USD")

    if role == "admin":
        currency_col, controls_spacer, country_col, manage_col, refresh_col = st.columns(
            [1.55, 3.45, 2.55, 1.20, 1.55],
            vertical_alignment="bottom",
        )
    else:
        currency_col, controls_spacer, refresh_col = st.columns(
            [1.55, 6.90, 1.55],
            vertical_alignment="bottom",
        )

    with currency_col:
        currency_mode = st.segmented_control(
            "CURRENCY",
            options=["USD", "Local Currency"],
            default=st.session_state.get("dashboard_currency_mode", "USD"),
            key="dashboard_currency_mode",
        ) or "USD"

    full_data = get_dashboard_dataset(currency_mode)
    if full_data.empty:
        st.warning("The consolidated file is empty.")
        return
    missing = validate_dashboard_columns(full_data)
    if missing:
        st.error("The dashboard cannot be built because the following columns are missing:")
        st.code(", ".join(missing))
        return

    country_column = find_column(full_data, "COUNTRY")
    all_countries = sorted(v for v in full_data[country_column].astype(str).map(clean_text).unique().tolist() if v)

    if role == "admin":
        with country_col:
            selected_countries = st.multiselect(
                "COUNTRY",
                options=all_countries,
                key="filter_country",
            )
        with manage_col:
            st.markdown('<div class="control-label-spacer"></div>', unsafe_allow_html=True)
            with st.popover("MANAGE DATA", use_container_width=True):
                st.caption("Remove all uploaded data for a country (originals, ML, USD, consolidated).")
                country_to_remove = st.selectbox(
                    "COUNTRY TO REMOVE",
                    options=all_countries,
                    key="manage_data_country",
                )
                confirm_remove = st.checkbox(
                    "I UNDERSTAND THIS CANNOT BE UNDONE",
                    key="manage_data_confirm",
                )
                if st.button(
                    "REMOVE COUNTRY DATA",
                    key="manage_data_remove_btn",
                    disabled=not confirm_remove,
                ):
                    remove_country_data(country_to_remove)
                    st.success(f"Data for {country_to_remove.title()} was removed.")
                    st.rerun()
        with refresh_col:
            st.markdown('<div class="control-label-spacer"></div>', unsafe_allow_html=True)
            render_refresh_button()
        if not selected_countries:
            selected_countries = all_countries
    else:
        with refresh_col:
            st.markdown('<div class="control-label-spacer"></div>', unsafe_allow_html=True)
            render_refresh_button()
        selected_countries = [country for country in all_countries if country.upper() == assigned_country.upper()]
        if not selected_countries:
            selected_countries = [assigned_country]

    if currency_mode == "Local Currency" and role == "admin" and len(selected_countries) > 1:
        st.warning("⚠ Local currency values from multiple countries are being added.")

    secured_data = full_data[full_data[country_column].astype(str).str.upper().isin(
        [c.upper() for c in selected_countries]
    )]
    if secured_data.empty:
        st.warning("No data available for the current user or selected country.")
        return

    current_filters = {
        "COUNTRY": selected_countries,
        "BU": st.session_state.get("filter_bu", []),
        "GENDER": st.session_state.get("filter_gender", []),
        "OCCASION": st.session_state.get("filter_occasion", []),
        "CATEGORY": st.session_state.get("filter_category", []),
    }
    cascade_source = secured_data
    filtered_data = filter_data_by_values(
        cascade_source,
        {
            "BU": current_filters["BU"],
            "GENDER": current_filters["GENDER"],
            "OCCASION": current_filters["OCCASION"],
            "CATEGORY": current_filters["CATEGORY"],
        },
    )

    currency_label = "USD" if currency_mode == "USD" else "Local Currency"
    render_metric_block("GENERAL TOTALS", calculate_dashboard_metrics(secured_data), currency_label)
    render_metric_block(
        "SUB TOTALS",
        calculate_dashboard_metrics(filtered_data if not filtered_data.empty else secured_data),
        currency_label,
    )

    st.markdown('<div class="dashboard-section-title">CATEGORY PLAN</div>', unsafe_allow_html=True)
    control_cols = st.columns([2.8, 1.3, 1.4, 1.5, 1.5], vertical_alignment="bottom")
    selections = {"COUNTRY": selected_countries}
    with control_cols[0]:
        hierarchy_level = st.segmented_control(
            "HIERARCHY LEVEL", options=HIERARCHY_LEVELS,
            default=st.session_state.get("dashboard_hierarchy_level", "BU"),
            key="dashboard_hierarchy_level",
        ) or "BU"
    for index, (label, column, key) in enumerate(
        [("BU", "BU", "filter_bu"), ("GENDER", "GENDER", "filter_gender"),
         ("OCCASION", "OCCASION", "filter_occasion"),
         ("CATEGORY", "CATEGORY", "filter_category")],
        start=1,
    ):
        with control_cols[index]:
            selections[column] = make_filter_multiselect(label, column, secured_data, selections, key)

    if filtered_data.empty:
        st.warning("No detail rows match the selected filters.")
        return

    vat_rate = vat_rate_for_data(filtered_data)
    table_data = build_hierarchy_summary(filtered_data, hierarchy_level, vat_rate)
    group_columns = [c for c in HIERARCHY_LEVELS[: HIERARCHY_LEVELS.index(hierarchy_level) + 1] if c in table_data.columns]
    dimension_columns = {"KEY", "BU", "GENDER", "OCCASION", "CATEGORY", "COUNTRY"}
    all_table_columns = [
        c for c in FULL_TABLE_ORDER
        if c is not None
        and c in table_data.columns
        and c not in OMIT_COLUMNS
        and c not in dimension_columns
        and not is_separator_column(c)
    ]
    default_columns = [c for c in DEFAULT_TABLE_ORDER if c is not None and c in all_table_columns]
    table_key = "table_column_filter"
    if table_key not in st.session_state:
        st.session_state[table_key] = default_columns

    visible_columns = st.multiselect(
        "TABLE COLUMNS",
        options=all_table_columns,
        key=table_key,
    )

    visible_columns = [
        column for column in visible_columns
        if column in all_table_columns
    ]

    if not visible_columns:
        visible_columns = default_columns if default_columns else all_table_columns

    metric_order = insert_separator_columns(visible_columns, FULL_TABLE_ORDER)
    display_order = group_columns + [c for c in metric_order if c not in dimension_columns]
    display_df = table_data.reindex(columns=[c for c in display_order if c in table_data.columns or is_separator_column(c)])
    for column in display_df.columns:
        if is_separator_column(column):
            display_df[column] = ""

    st.dataframe(
        style_dashboard_table(display_df),
        height=620,
        hide_index=True,
        width="stretch",
        column_config=build_column_config(display_df.columns),
    )

    st.markdown('<div class="dashboard-section-title">PERFORMANCE CHARTS</div>', unsafe_allow_html=True)
    st.markdown(
        f'<div class="dashboard-section-subtitle">Share of business and combined KPI trends by {hierarchy_level}, using filtered data.</div>',
        unsafe_allow_html=True,
    )
    chart_col1, chart_col2 = st.columns(2)
    with chart_col1:
        render_share_chart(
            filtered_data, hierarchy_level, "TURNOVER VALUE 2027",
            f"Turnover 2027 - Share by {hierarchy_level}",
        )
    with chart_col2:
        render_share_chart(
            filtered_data, hierarchy_level, "PAIRS VALUE 2027",
            f"Volume 2027 - Share by {hierarchy_level}",
        )

    combined_col1, combined_col2 = st.columns(2)

    turnover_25 = numeric_sum(filtered_data, "TURNOVER VALUE 2025")
    turnover_26 = numeric_sum(filtered_data, "TURNOVER VALUE 2026")
    turnover_27 = numeric_sum(filtered_data, "TURNOVER VALUE 2027")
    mrg_25 = numeric_sum(filtered_data, "MRG VALUE 2025")
    mrg_26 = numeric_sum(filtered_data, "MRG VALUE 2026")
    mrg_27 = numeric_sum(filtered_data, "MRG VALUE 2027")

    margin_pct_25 = safe_divide(mrg_25, turnover_25)
    margin_pct_26 = safe_divide(mrg_26, turnover_26)
    margin_pct_27 = safe_divide(mrg_27, turnover_27)

    pairs_25 = numeric_sum(filtered_data, "PAIRS VALUE 2025")
    pairs_26 = numeric_sum(filtered_data, "PAIRS VALUE 2026")
    pairs_27 = numeric_sum(filtered_data, "PAIRS VALUE 2027")

    asp_25 = safe_divide(turnover_25, pairs_25)
    asp_26 = safe_divide(turnover_26, pairs_26)
    asp_27 = safe_divide(turnover_27, pairs_27)

    with combined_col1:
        render_combined_chart(
            filtered_data,
            title="Turnover & Margin %",
            bar_label="Turnover",
            bar_values=[turnover_25, turnover_26, turnover_27],
            line_label="Margin %",
            line_values=[margin_pct_25, margin_pct_26, margin_pct_27],
            line_is_percent=True,
        )
        turnover_margin_table = pd.DataFrame({
            "": ["TURNOVER", "MARGIN %"],
            "2025": [
                format_integer(turnover_25),
                format_percent(margin_pct_25, 0),
            ],
            "2026": [
                format_integer(turnover_26),
                format_percent(margin_pct_26, 0),
            ],
            "2027": [
                format_integer(turnover_27),
                format_percent(margin_pct_27, 0),
            ],
        })
        st.table(turnover_margin_table)

    with combined_col2:
        render_combined_chart(
            filtered_data,
            title="Volume & ASP",
            bar_label="Volume",
            bar_values=[pairs_25, pairs_26, pairs_27],
            line_label="ASP",
            line_values=[asp_25, asp_26, asp_27],
            line_is_percent=False,
        )
        volume_asp_table = pd.DataFrame({
            "": ["VOLUME", "ASP"],
            "2025": [
                format_integer(pairs_25),
                format_decimal(asp_25, 2),
            ],
            "2026": [
                format_integer(pairs_26),
                format_decimal(asp_26, 2),
            ],
            "2027": [
                format_integer(pairs_27),
                format_decimal(asp_27, 2),
            ],
        })
        st.table(volume_asp_table)

    download_col1, download_col2 = st.columns(2)

    if role == "admin":
        with download_col1:
            if GLOBAL_LOCAL_DATA_CSV.exists():
                local_export = read_global_csv(GLOBAL_LOCAL_DATA_CSV)
                buffer = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Consolidated ML"
                ws.append(list(local_export.columns))
                for _, row_values in local_export.iterrows():
                    ws.append(list(row_values))
                wb.save(buffer)
                wb.close()
                data_bytes = buffer.getvalue()
            else:
                data_bytes = None

            render_download_button(
                label="DOWNLOAD ML CONSOLIDATED",
                data_bytes=data_bytes,
                file_name="Category_Global_ML.xlsx",
                empty_caption="No local currency consolidated data available yet.",
            )

        with download_col2:
            data_bytes = GLOBAL_XLSX.read_bytes() if GLOBAL_XLSX.exists() else None
            render_download_button(
                label="DOWNLOAD USD CONSOLIDATED",
                data_bytes=data_bytes,
                file_name="Category_Global_USD.xlsx",
                empty_caption="No USD consolidated file available yet.",
            )
    else:
        country_ml_path = LOCAL_DIR / f"{assigned_country}.xlsx"
        country_usd_path = USD_DIR / f"{assigned_country}.xlsx"

        with download_col1:
            data_bytes = country_ml_path.read_bytes() if country_ml_path.exists() else None
            render_download_button(
                label="DOWNLOAD ML FILE",
                data_bytes=data_bytes,
                file_name=f"{assigned_country}_ML.xlsx",
                empty_caption="No local currency file available yet. Upload data first.",
            )

        with download_col2:
            data_bytes = country_usd_path.read_bytes() if country_usd_path.exists() else None
            render_download_button(
                label="DOWNLOAD USD FILE",
                data_bytes=data_bytes,
                file_name=f"{assigned_country}_USD.xlsx",
                empty_caption="No USD file available yet. Upload data first.",
            )

# ==========================================================
# ENTRYPOINT
# ==========================================================
st.session_state.setdefault("authenticated", False)
st.session_state.setdefault("page", "upload")
st.session_state.setdefault("dashboard_hierarchy_level", "BU")

def render_footer() -> None:
    st.markdown(
        """
        <div class="app-footer">
            © 2026 Category Plan  |  Merchandising Global  |  
            For assistance: <a href="mailto:ruben.villalon@bata.com">ruben.villalon@bata.com</a>
        </div>
        """,
        unsafe_allow_html=True,
    )

if st.session_state["authenticated"]:
    if st.session_state.get("page") == "dashboard":
        show_dashboard_page()
    else:
        show_upload_page()
else:
    show_login()

render_footer()